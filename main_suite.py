from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5 import QtCore
from PyQt5.QtGui import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import os, sys
import numpy as np
import pandas as pd
import pickle
import subprocess
import sys
import win32com.client as win32
import time
import re
import shutil # 파일 복사용 모듈
import Modules.NFS_DNA as NFS_DNA
from time import sleep


# QtDesigner로 만든 UI 파일을 로딩
form_entry = uic.loadUiType("GUI/EntryForm.ui")[0]
form_main_suite = uic.loadUiType('GUI/MainSuiteForm.ui')[0]


class DataDNAIdentification:
    """
    채취 한 텀에서 얻은 데이터와 해당 데이터의 저장 정보를 보관하는 클래스

    Attributes
    ----------
    location_save : str
        해당 텀의 데이터를 보관하는 폴더의 위치
    analyst : str
        해당 텀에 담당자
    date : str
        채취 일자
    df_evidence : pandas.DataFrame
        NFIS 상에서 다운로드한 감정처리부 데이터와 감정처리부를 가공한 데이터, 실험 여부 등을 감정물 별로 저장하는 데이터프레임
    nfis_loaded : bool
        NFIS 파일을 df_evidence에 입력했는지 여부
    list_tag : list
        감정물의 기본 분류 (e.g REF, LCN)
    path_tomato : string
        해당 프로젝트의 tomato 파일 경로
    path_totalsheet : string
        해당 프로젝트의 total_sheet 파일 경로
    path_resamplesheet : string
        해당 프로젝트의 RESAMPLING 파일 경로
    path_picture : string
        해당 프로젝트의 감정물 사진이 보관된 폴더 경로
    combined_result : NFS_DNA.CombinedResult
        Tomato 파일에서 CombinedResult 탭의 정보를 parsing한
    Methods
    -------
    change_path(path)
        EntryForm에서 ddi를 불러들일 때 ddi가 위치했던 경로로 각종 파일의 경로를 변경(폴더가 생성 위치에서 옮겨졌을 때 대비)
    get_defaultname()
        파일 이름에 쓰일 기본 이름(날짜+담당자(+ 특이사항))을 반환한다.
    """

    def __init__(self, location_save, analyst, date):
        """
        Parameters
        ----------
        location_save : str
             해당 텀의 데이터를 보관하는 폴더의 위치
        analyst : str
            해당 텀에 담당자
        date : str
            채취 일자
        """

        self.location_save = location_save
        self.analyst = analyst
        self.date = date
        self.df_evidence = pd.DataFrame()   # NFIS감정처리부 파일을 DataFrame으로 저장
        self.df_report = pd.DataFrame() # 감정서 작성용 DataFrame
        self.nfis_loaded = False
        self.list_tag = ['LCN', 'MF', 'REF'] # 기본분류
        self.path_tomato = location_save + '/' + date + '-' + analyst + '-Tomato-TOTAL.xlsm'
        self.path_tomato_y23 = location_save + '/' + date + '-' + analyst + '-Tomato-Y23.xlsm'
        self.path_totalsheet = location_save + '/Sheets/' + date + '-' + analyst + '-' + 'TOTAL.xlsm'
        self.path_resamplesheet = location_save + '/Sheets/' + date + '-' + analyst + '-' + 'RESAMPLING.xlsm'
        self.path_picture = location_save + '/감정물사진/'
        self.combined_result = NFS_DNA.CombinedResult(kit="GF/PPF")
        self.combined_result_y23 = NFS_DNA.CombinedResult(kit="Y23")

    def change_path(self, path):
        """
        EntryForm에서 ddi를 불러들일 때 ddi가 위치했던 경로로 각종 파일의 경로를 변경(폴더가 생성 위치에서 옮겨졌을 때 대비)

        Parameter
        ---------
        path : str
            변경할 폴더 위치
        """
        self.location_save = path
        self.path_tomato = path + '/' + self.date + '-' + self.analyst + '-Tomato-TOTAL.xlsm'
        self.path_totalsheet = path + '/Sheets/' + self.date + '-' + self.analyst + '-' + 'TOTAL.xlsm'
        self.path_resamplesheet = path + '/Sheets/' + self.date + '-' + self.analyst + '-' + 'RESAMPLING.xlsm'
        self.path_picture = path + '/감정물사진/'

    def get_defaultname(self):
        """
        파일 이름에 쓰일 기본 이름(날짜+담당자)을 반환한다.

        Returns
        -------
        str
            기본 파일명(채취날짜_담당자))
        """
        defaultname = '%s-%s' % (self.date, self.analyst)
        return defaultname


class EntryForm(QDialog, form_entry):
    """
    프로그램 시작 시 새로 프로젝트 폴더를 생성하거나 기존 폴더의 데이터를 불러올 도입부의 GUI 구현하는 클래스

    새로 폴더를 생성 시 입력받은 데이터를 토대로 DataDNAIdentification 객체를 생성하여 MainSuiteForm 객체에 전달한다.
    기존 폴더를 선택 시 선택된 폴더 내의 DataDNAIdentification 클래스의 pickle 파일을 unpickle해서 MainSuiteForm 객체에 전달한다.

    Inheritance
    ---------
    QDialog :
        PyQt의 다이얼로그 창 객체
    form_entry :
        QtDesigner로 작성된 도입부의 GUI 코드

    Attributes
    ----------
    GUI_main_suite : MainSuiteForm
        데이터를 전달하고 실행시킬 MainSuiteForm의 객체

    Methods
    -------
    set_line_texts(location_save="", analyst="", date=QtCore.QDate.currentDate())
        저장위치, 담당자, 채취날짜를 인자로 받아 각각 해당하는 QLineEdit, QDateEdit 객체의 Text 속성에 할당한다.
    set_line_ro(condition=True)
        도입부 GUI의 QLineEdit, QDateEdit 객체들의 읽기전용 속성을 전환한다.

    click_btn_new()
        QPushButton 객체인 btn_new의 클릭 이벤트. 새로 폴더를 생성하기 위한 정보를 입력 받기 위해 QLineEdit, QDateEdit 객체들의 읽기전용 속성을 해제한다.
    click_btn_ok()
        QPushButton 객체인 btn_ok의 클릭 이벤트. 입력 받은 값을 토대로 폴더와 DataDNAIdentification 객체를 생성하고 GUI_main_suite에 인자로 넘긴 후 메인 GUI를 활성화한다.
    click_btn_load()
        QPushButton 객체인 btn_load의 클릭 이벤트. 지정한 폴더 내 DataDNAIdentification 객체 pickle 데이터를 읽고 GUI_main_suite에 인자로 넘긴 후 메인 GUI를 활성화한다.
    click_btn_exit()
        QPushButton 객체인 btn_exit의 클릭 이벤트. 프로그램을 종료한다.
    """

    def __init__(self):
        """
        EntryForm 클래스의 생성자. Setting.ini 파일 내 기본값을 읽고 QLineEdit 객체들에 반영한다.
        """

        super().__init__()
        self.setupUi(self)
        self.root = os.getcwd()
        with open(os.getcwd() + '/Settings/Settings.ini', mode='r') as readfile_setting:
            location_save = readfile_setting.readline().rstrip('\n').split('=')[1]
            analyst = readfile_setting.readline().rstrip('\n').split('=')[1]
            readfile_setting.close()
        self.set_line_texts(location_save=location_save, analyst=analyst,
                            date=QtCore.QDate.currentDate())
        self.GUI_main_suite = None

    def set_line_texts(self, location_save="", analyst="", date=QtCore.QDate.currentDate()):
        """
        저장위치, 담당자, 채취날짜, 특이사항을 인자로 받아 각각 해당하는 QLineEdit 객체의 Text 속성에 할당한다.

        Parameters
        ----------
        location_save : str
             해당 텀의 데이터를 보관하는 폴더의 위치
        analyst : str
            해당 텀에 담당자
        date : QtCore.QDate
            채취 일자 (default = 오늘 날짜)
        """

        self.line_savelocation.setText(location_save)
        self.line_analyst.setText(analyst)
        self.line_date.setDate(date)

    def set_line_ro(self, condition=True):
        """
        도입부 GUI의 QLineEdit, QDateEdit 객체들의 읽기전용 속성을 전환한다.

        Parameters
        ----------
        condition : bool, optional
            편집가능한 GUI 객체들의 ReadOnly 속성의 변경값
        """

        self.line_analyst.setReadOnly(condition)
        self.line_date.setReadOnly(condition)

    def click_btn_new(self):
        """
        QPushButton 객체인 btn_new의 클릭 이벤트. 새로 폴더를 생성하기 위한 정보를 입력 받기 위해 QLineEdit, QDateEdit 객체들의 읽기전용 속성을 해제한다.
        """

        QMessageBox.information(self, "Notice", "Input new profile and press OK button.")
        self.set_line_ro(False)

    def click_btn_ok(self):
        """
        QPushButton 객체인 btn_ok의 클릭 이벤트. 입력 받은 값을 토대로 폴더와 DataDNAIdentification 객체를 생성하고 GUI_main_suite에 인자로 넘긴 후 메인 GUI를 활성화한다.

        '채취날짜_감정인'을 이름으로 가지는 폴더를 생성하고 실험 결과들을 모아두는데 필요한 하위폴더를 생성한다.
        저장위치, 채취날짜 감정인을 인자로 DataDNAIdentification 객체를 생성한다
        해당 객체를 GUI_main_suite에 인자로 넘겨준 후 메인 GUI를 활성화하고 현재 GUI 객체를 닫는다.

        """

        new_dir = '%s/%s_%s' % (self.line_savelocation.text(),
                                self.line_date.date().toString('yyyyMMdd'),
                                self.line_analyst.text())
        os.chdir(self.line_savelocation.text())
        if not os.path.exists(new_dir):
            os.mkdir(new_dir)
            os.chdir(new_dir)
            os.mkdir('Downloaded')  # Import한 파일들을 복사해서 저장할 위치
            os.mkdir('Sheets') # 샘플시트들을 저장할 위치
            os.mkdir('RT') # RT 관련 파일들을 저장할 위치
            os.mkdir('DATA') # 리딩 후 생성된 데이터를 보관할 위치
            os.mkdir('DB')  # DB 관련된 데이터를 보관할 위치
            os.mkdir('ETC') # 기타 생성된 시트 및 데이터를 보관할 위치
            os.mkdir('Reports')
            os.mkdir('감정물사진')
            # 하위 폴더가 추가로 필요할 경우 이곳에 추가
            shutil.copyfile(self.root + '/Form/form_Tomato_Tools_18.10_customized.xlsm', new_dir + '/' + self.line_date.date().toString('yyyyMMdd') + "-" + self.line_analyst.text() + "-Tomato-TOTAL.xlsm")
        else:
            QMessageBox.information(self, 'Error', 'Same folder exists')
            return
        os.chdir(self.root) #current working directory를 실행파일이 존재하는 디렉토리로 변경
        ddi_new = DataDNAIdentification(location_save=os.path.realpath(new_dir), analyst=self.line_analyst.text(),
                                        date=self.line_date.date().toString('yyyyMMdd'))
        self.GUI_main_suite = MainSuiteForm(ddi_new)
        self.GUI_main_suite.show()
        self.close()

    def click_btn_load(self):
        """
        QPushButton 객체인 btn_load의 클릭 이벤트. 지정한 폴더 내 DataDNAIdentification 객체 pickle 데이터를 읽고 GUI_main_suite에 인자로 넘긴 후 메인 GUI를 활성화한다.

        pickling된 DataDNAIdentifiacitoin 객체를 불러오기 위해 pickle 라이브러리를 사용.
        unpickle한 DataDNAIdentifiacitoin 객체를 GUI_main_suite에 인자로 넘겨준 후 메인 GUI를 활성화하고 현재 GUI 객체를 닫는다.

        Raises
        ------
        UnpicklingError
            pickling된 DataDNAIdentification 객체를 unpickle하는 작업이 실패할 경우
        IOError
            잘못된 폴더 또는 파일명(e.g. pickle파일이 없는 폴더, 폴더 선택 x)을 처리할 경우
        """

        location_load = QFileDialog.getExistingDirectory(self, 'Open folder', self.line_savelocation.text())
        try:
            with open('%s/%s' % (location_load, 'DataDNAIdentification.pickle'), 'rb') as f:
                ddi_load = pickle.load(f)
        except pickle.UnpicklingError:
            QMessageBox.information(self, 'Unpickling Error', 'Inappropriate .pickle file')
        except IOError:
            QMessageBox.information(self, 'I/O Error', 'Inappropriate folder selected')
        else:
            ddi_load.change_path(location_load)
            self.GUI_main_suite = MainSuiteForm(ddi_load)
            self.GUI_main_suite.show()
            self.close()

    def click_btn_exit(self):
        """
        QPushButton 객체인 btn_load의 클릭 이벤트. 프로그램을 종료한다.
        """
        sys.exit()
    

class MainSuiteForm(QMainWindow, form_main_suite):
    """
    프로그램의 주요 기능 및 인터페이스 구현

    @ Info. tab
        기본 정보, 진행 상황, 기타 수치를 나타내고, 업무와 관련된 외부 프로그램을 바로 실행할 수 있는 기능 탑재
    @ Sheets tab
        Sample sheets : NFIS 파일을 읽고 각각의 감정물을 유형 별로 나누고 ddi_present의 df_evidence에 저장
        Generate RT sheets : 토탈샘플시트 파일로부터 RT 시트를 생성
        Import RT data : RT-PCR 결과값 파일을 토탈샘플시트에 반영한다
    @ Data Tab
        export_to_barcode : ddi_present의 df_evidence의 데이터를 form_barcode.xls에 복사한다
        onsite_request: 업무분장 NFIS 파일을 입력받아 소내의뢰 시트와 증거물에 붙힐 라벨을 생성한다

    Inheritance
    -----------
    QMainWindow :
        PyQt의 메인 윈도우 객체
    form_main_suite :
        QtDesigner로 작성된 메인부의 GUI 코드

    Attributes
    ----------
    ddi_present : DataDNAIdentification
        NFIS 파일의 데이터와 실험결과를 모아서 정리할 DataDNAIdentification 객체
    root : str
        해당 프로그램의 루트 폴더의 위치
    exapp : dict
        key : 외부 프로그램 이름, item : 해당 외부 프로그램의 위치

    Methods
    -------
    @ internal function
        run_external_app(app)
            인자로 받은 이름에 해당하는 외부 프로그램을 실행
        save()
            ddi_present 객체를 pickle하여 DataDNAIdentification.pickle 파일에 저장
        sort_by_serial(df)
            입력받은 dataframe을 증거물 번호를 기준으로 natural sort
        import_file(extension="", copy_needed=True)
            pyqt 파일 다이얼로그 상에서 파일을 선택하고 해당 파일의 경로를 반환환
        xls_to_dataframe(file_input = "", column = True)
            NFIS에서 받은 엑셀 파일을 Dataframe 객체로 전환해서 반환
        update_df_sample(df, target_list, tag)
            샘플시트 작성 시 target_list에 배정된 감정물에 입력된 분류명을 df에 기록
        move_all_item(from_list, to_list)
            한 리스트 위젯에 있는 모든 아이템의 내용을 다른 리스트 위젯으로 이동
        move_items(from_list, to_list, selected_item)
            한 리스트 위젯에 있는 선택한 아이템의 내용 다른 리스트 위젯으로 이동
        idx_to_wellname(idx):
            8x12 wells plate를 기준으로 입력된 index 번호에 해당하는 well의 명칭(e.g. 0->A1)을 반환
        wellname_to_idx(wellname):
            8x12 wells plate를 기준으로 입력된 wellname에 해당하는 index 번호(e.g. A1->0)를 반환
        click_btn_open_savelocation()
            btn_open_savelocation의 클릭 이벤트. 해당 이벤트를 호출한 버튼의 이름에 할당된 저장 폴더를 연다.
    @ info tab
        set_line_texts(location_save="", analyst="", date=QtCore.QDate.currentDate().toString('yyyyMMdd'))
            현재 작업의 저장위치, 담당자, 채취날짜를 인자로 받아 각각 해당하는 QLineEdit, QDateEdit 객체의 Text 속성에 할당한다.
        search_table(keyword, table)
            해당 QTableWidget에서 keyword를 가진 아이템을 찾아 커서를 움직인다.
        click_btn_search_info()
            btn_search_info의 클릭 이벤트, line_search_info의 키워드를 table_info에서 찾아 커서를 움직인다.
        click_btn_NFIS_login():
            btn_NFIS_login의 클릭 이벤트. 외부 프로그램 NFIS_login을 실행한다.
        click_btn_NFIS_revision_helper(self):
            btn_NFIS_revision_helper의 클릭 이벤트. 외부 프로그램 NFIS_revision_helper를 실행한다.
        click_btn_NFIS_tomato():
            btn_NFIS_tomato의 클릭 이벤트. 해당 프로젝트의 Tomato 엑셀 파일을 연다.
        click_btn_total_sheet():
            btn_total_sheet의 클릭 이벤트. 해당 프로젝트의 total_sheet 엑셀 파일을 연다.
    @ sheet tab - sample sheet
        load_samplesheets()
            현재 ddi_present에 저장된 데이터에 따라 Sheet탭의 리스트와 콤보박스를 업데이트한다
        generate_samplesheets(worksheet, df, filename, row_start,  control = False, blank = False, ladder = False)
            해당 워크시트에 ddi_present 객체의 값을 입력한 후 지정된 파일 이름으로 저장한다.
        change_combo_category(item)
            combo_category의 변경 이벤트. combo_category에서 아이템을 선택했을 때 해당 분류에 속하는 증거물의 목록을 QListWidget객체에 반영한다.
        update_list_count(qListWidget, QLabel)
            리스트의 변경된 item 개수를 QLabel에 반영
        click_btn_import_modified_sample()
            btn_import_modified_sample의 클릭 이벤트. 채취 후 수정한 NFIS파일을 읽고 ddi_present 객체에 저장한다. 그 후 증거물 목록을 list_sample_all에 반영한다.
        click_btn_add_category()
            btn_add_category의 클릭 이벤트. 새로운 증거물 분류를 combo_category 추가한다.
        click_btn_remove_category()
            btn_remove_category의 클릭 이벤트. 현재 선택된 증거물 분류를 combo_category에서 제거한다.
        click_btn_move_all()
            btn_move_all의 클릭 이벤트. list_sample_all의 모든 내용을 list_sample_partial으로 옮긴다.
        click_btn_remove_all()
            btn_remove_all의 클릭 이벤트. list_sample_partial의 모든 내용을 list_sample_all으로 옮긴다.
        click_btn_move()
            btn_move의 클릭 이벤트. 선택된 아이템을 list_sample_all에서 list_sample_partial으로 옮긴다.
        click_btn_remove()
            btn_remove의 클릭 이벤트. 선택된 아이템을 list_sample_partial에서 list_sample_all으로 옮긴다.
        click_btn_generate_samplesheets()
            btn_generate_samplesheets의 클릭 이벤트. ddi_present 저장된 데이터를 분류대로 나눠 샘플시트를 생성한다.
        click_btn_generate_totalsheet()
            btn_generate_totalsheet의 클릭 이벤트. ddi_present 저장된 데이터를 하나의 샘플시트로 생성한다.
    @ sheet tab - RT
        click_btn_generate_RT_sheet_from_total()
            btn_generate_RT_sheet_from_total 버튼의 클릭 이벤트. totalsheet 엑셀 파일의 TOTAL 시트에서 TYPE이 LCN, REF인 것만 추출하여 RT import 파일을 작성한다.
        click_btn_import_RT()
            btn_Import_RT 버튼의 클릭 이벤트. RT 실험 결과 파일의 경로를 입력받는다. 그리고 해당 파일의 RT 실험 결과를 증거물 토탈샘플시트 파일에 복사한다.
    @ Report tab
        load_reportsheets(self)
            감정서 데이터프레임 내의 접수번호를 리스트로 만들고 combo_report_cases에 반영한다
         change_combo_report_cases(self, item)
            선택된 사건번호에 따라 감정서 테이블을 갱신한다.
         click_btn_report_next(self)
            combo_report_cases의 다음 item 선택
         click_btn_load_tomato(self)
            Tomato 엑셀 파일의 combined_result 탭에서 DNA profile 데이터를 NFS_DNA 클래스 상에 불러온다. Y23 Tomamto 파일이 있다면 해당 파일의 데이터도 불러온다.
         update_table_report(self, number_case)
            table_report에 ddi_present의 df_report값을 입력한다.\
         cellchange_table_report(self, row, col)
            증거물 테이블의 내용이 변경되면 변경된 내용을 감정서 데이터프레임에 반영한다
         load_image(self, path)
            입력받은 경로의 이미지를 label 객체에 띄운다
         load_list_images(self, case_number)
            해당 사건번호을 파일이름에 포함하는 이미지의 이름을 찾아서 list_images에 반영한다
         click_list_picture_item(self, item)
            list_images에서 클릭된 아이템을 파일이름으로 가지는 이미지를 label 객체에 띄운다
         click_btn_generate_report(self)
            생성할 감정서 종류와 선택된 사건번호의 데이터를 토대로 해당 감정서를 작성한다.
         generate_report(self, num_case, type_report)
            선택된 사건번호를 생성할 감정서 종류에 맞춰 감정서 hwp 파일을 생성한다.
         click_btn_export_barcode(self)
            ddi_present의 df_evidence의 데이터를 form_barcode.xls에 복사한다
         click_btn_onsite_request(self)
            업무분장 NFIS 파일을 입력받아 소내의뢰 시트와 증거물에 붙힐 라벨을 생성한다
    @ Data Tab
        click_btn_export_to_barcode() : ddi_present의 df_evidence의 데이터를 form_barcode.xls에 복사한다
        click_btn_onsite_request(): 업무분장 NFIS 파일을 입력받아 소내의뢰 시트와 증거물에 붙힐 라벨을 생성한다
    """

    def __init__(self, ddi):
        """
        MainSuiteForm의 생성자. DataDNAIdentification 객체를 인자로 받고 프로그램에 쓰이는 기본 설정을 처리한다.

        Parameter
        ---------
        ddi ; DataDNAIdentification
            프로그램 실행 시 불러온 작업을 수행할 DataDNAIdentification 객체
        """

        super().__init__()
        self.setupUi(self)
        self.ddi_present = ddi
        self.root = os.path.dirname(os.path.abspath(__file__))
        self.set_line_texts(location_save=self.ddi_present.location_save, analyst=self.ddi_present.analyst, date=self.ddi_present.date)
        if ddi.nfis_loaded == True: #기존에 읽어드린 NFIS 파일이 있다면 해당 DataFrame의 내용을 GUI에 반영하고, 아니면 해당 탭을 비활성화
            self.load_samplesheets()
            self.load_resamplesheets()
            self.load_reportsheets()
        else:
            self.tabWidget.setTabEnabled(2, False)  # Resample tab 비활성화
            self.tabWidget.setTabEnabled(3, False)   # Report tab 비활성화
            self.tabWidget.setTabEnabled(4, False)   # Data tab 비활성화
        self.exapp = {}
        with open(self.root+"/Settings/External_apps.ini", mode='r') as readfile_exapp: # 외부 프로그램의 위치를 딕셔너리에 저장
            for lines in readfile_exapp:
                line_sep = lines.split('=')
                self.exapp[line_sep[0]] = line_sep[1]
        self.update_info_table()
        self.dispatch_excel = win32.Dispatch('Excel.Application')  # 엑셀을 다루기 위해 사용할 핸들러
        self.path_form_report = {'ND' : '/Form/form_report_ND.hwp',
                                 '부검' : '/Form/form_report_D.hwp',
                                 '피해자 일치' : '/Form/form_report_V-match.hwp',
                                 'ND w/ 피해자 일치' : '/Form/form_report_V-match+ND.hwp',
                                 'ND w/ 피해자 불일치' : '/Form/form_report_V-nonmatch+ND.hwp',
                                 'Complicate' : '/Form/form_report_Complicate.hwp',
                                 '혼합형' : '/Form/form_report_MX.hwp',
                                 '피의자 일치' : '/Form/form_report_S-match.hwp',
                                 '피의자 불일치' : '/Form/form_report_S-nonmatch.hwp',
                                 '친자관계 일치' : '/Form/form_report_Parent-Child.hwp',
                                 '친자관계 일치(부검)' : '/Form/form_report_Parent-Child_D.hwp',
                                 'C 검출(검색결과 X)' : '/Form/form_report_C.hwp',
                                 'C 검출 w/ ND(검색결과 X)' : '/Form/form_report_C+ND.hwp',
                                 'C 검출 w/ 피해자 불일치(검색결과 X)' : '/Form/form_report_C+V-nonmatch.hwp',
                                 'C 검출 w/ 피해자 불일치, ND(검색결과 X)' : '/Form/form_report_C+ND+V-nonmatch.hwp',
                                 'C 검출 w/ 피해자 일치(검색결과 X)': '/Form/form_report_C+V-match.hwp',
                                 'C 검출 w/ 피해자 일치, ND(검색결과 X)': '/Form/form_report_C+ND+V-match.hwp'
                                 }
        self.save()

    def closeEvent(self, event):    # 엑셀을 다루기 사용했던 Win32com.client를 닫아주고 df_evidence를 자동저장하기 위해 QWidget의 closeEvent를 오버라이드.
        self.save()
        self.dispatch_excel.Quit()
        event.accept()

    # internal function
    def run_external_app(self, app):
        """
        인자로 받은 이름에 해당하는 외부 프로그램을 실행

        Parameters
        -----------
        app : str
            exapp에 저장된 실행할 외부 프로그램의 key
        """

        subprocess.SW_HIDE = 1
        r = subprocess.Popen(self.root + self.exapp[app], shell=True)
        if r == 1:
             print('running error')

    def save(self):
        """
        ddi_present 객체를 pickle하여 DataDNAIdentification.pickle 파일에 저장
        """

        with open(self.ddi_present.location_save + '/DataDNAIdentification.pickle', 'wb') as f:
            pickle.dump(self.ddi_present, f, pickle.HIGHEST_PROTOCOL)

    def sort_by_serial(self, df):
        """
        입력받은 증거물 데이터프레임을 증거물번호를 기준으로 natural sort한다.

        우선 증거물 데이터 프레임을 증거물번호를 기준으로 일반적인 방식을 사용해 정렬한다.
        정렬된 증거물번호를 리스트로 만들고 해당 리스트를 증거물 번호를 구성하는 세 종류의 숫자(e.g 2019-D-1234-5 => 2019, 1234, 5)를 첫번째, 두번째, 세번째 기준으로 삼아 정렬한다.
        해당 리스트를 enumerate 객체로 만들어 인덱스를 부여하고 인덱스-증거물번호로 이뤄진 딕셔너리 rank를 생성한다.
        rank를 데이터프레임에 병합한 후 rank를 기준으로 데이터프레임을 정렬한다.
        사용된 rank 데이터를 제거하고 정렬된 순서에 맞추어 새로 인덱스를 부여한다.

        Parameters
        ---------
        df : DataFrame
            정렬할 증거물 데이터프레임
        """

        df.sort_values('증거물번호', inplace=True)
        list_serial = list(df['증거물번호'])
        p = re.compile('\d+')
        sorted_serial = sorted(list_serial, key = lambda x: (int(p.findall(x)[0]), int(p.findall(x)[1]), int(p.findall(x)[2]), int(p.findall(x)[3])) if len(p.findall(x))>3 else (int(p.findall(x)[0]), int(p.findall(x)[1]),int(p.findall(x)[2]), 0))
        # 2019-D-1234-5 => 2019, 1234, 5
        sorted_serial = enumerate(sorted_serial)    # rank로 쓰일 인덱스 부여를 위해 enumerate 객체로 전환
        rank = {k:v for v,k in sorted_serial}
        df['new_index'] = df['증거물번호'].apply(lambda x:rank[x])
        df.sort_values('new_index', inplace=True)
        df.drop('new_index', axis=1, inplace=True)
        df.reset_index(inplace=True)

    def import_file(self, extension="", copy_needed=True, title = "Open File"):
        """
        pyqt 파일 다이얼로그 상에서 파일을 선택하고 해당 파일의 경로를 반환환다.

        copy_needed가 True면 해당 경로의 파일을 지정된 위치에 복사한다.

        Parameters
        -----------
        extension : str, optional
            import할 파일의 확장자명
        copy_needed : bool, optional
            타경로에서 import할 경우 해당 파일의 복사본을 작업 폴더의 Downloaded 폴더에 복사 할 것인지 여부 (default = True)

        Return
        ------
        str
            pyqt 다이얼로그에서 입력받은 파일 경로
        """

        filename_import = QFileDialog.getOpenFileName(self, title, self.ddi_present.location_save, extension)
        if filename_import[0] is "":
            QMessageBox.information(self, "Error", "File name required.")
            return
        if copy_needed is True:
            shutil.copyfile(filename_import[0], self.ddi_present.location_save + '/Downloaded/'
                            + filename_import[0].split('/')[-1])
        return filename_import[0]

    def xls_to_dataframe(self, file_input = "", column = True):
        """
        NFIS에서 받은 엑셀 파일을 Dataframe 객체로 전환해서 반환

        Parameters
        -----------
        file_input : str
            openpyxl 라이브러리로 작업할 NFIS 파일의 경로
        column : bool
            column 헤더의 존재 여부

        Returns
        --------
        DataFrame
            NFIS 파일의 내용을 DataFrame으로 변환한 객체
        """

        wb = load_workbook(file_input, data_only=True, read_only=False)
        ws = wb.active
        data = ws.values
        if column is True:
            cols = next(data)
            return pd.DataFrame(data, columns=cols)
        else:
            return pd.DataFrame(data)

    def update_df_sample(self, df, target_list, tag):
        """
        샘플시트 작성 시 target_list에 배정된 감정물에 입력된 분류명을 증거물 데이터프레임에 기록

        target_list의 item의 text에서 인덱스를 추출해서 증거물 데이터프레임의 해당 인덱스의 분류에 tag를 입력

        Parameters
        ----------
        df : DataFrame
            편집할 증거물 데이터프레임
        target_list : QListWidget
            tag로 나눠진 QListWidget
        tag : str
            입력할 분류명
        """

        for row_number in range(target_list.count()):
            ind = int(target_list.item(row_number).text().split(' ')[0]) - 1
            df.loc[ind, '분류'] = tag

    def move_all_item(self, from_list, to_list):
        """
        한 리스트 위젯에 있는 모든 아이템의 내용을 다른 리스트 위젯으로 이동

        Parameters
        ----------
        from_list : QListWidget
             넘겨줄 내용을 담은 QListWidget
        to_list : QListWidget
             내용을 받을 QListWidget
        """

        for row_number in range(from_list.count()):
            to_list.addItem(from_list.item(row_number).text())
        from_list.clear()

    def move_items(self, from_list, to_list, selected_item):
        """
        한 리스트 위젯에 있는 선택한 아이템의 내용 다른 리스트 위젯으로 이동

        Parameters
        ----------
        from_list: QListWidget
            넘겨줄 내용을 담은 QListWidget
        to_list: QListWidget
            내용을 받을 QListWidget
        selected_item: List
            from_list에서 선택된 아이템(들)의 list
        """

        to_list.addItems(item.text() for item in selected_item)
        for item in selected_item:
            from_list.takeItem(from_list.row(item))

    def idx_to_wellname(self, idx):
        """
        8x12 wells plate를 기준으로 입력된 index 번호에 해당하는 well의 명칭(e.g. 0->A1)을 반환

        Parameters
        ----------
        idx: int
            변환할 인덱스
        """

        col, row = divmod(idx,8)
        wellname = 'ABCDEFGH'[row] + str(col+1)
        return wellname # 1줄로 : return str(idx//8+1)+'HABCDEFG'[idx%8]

    def wellname_to_idx(self, wellname):
        """
        8x12 wells plate를 기준으로 입력된 wellname에 해당하는 index 번호(e.g. A1->0)을 반환

        Parameters
        ----------
        wellname : str
            변환할 wellname
        """

        alphabet = wellname[0]
        number = int(wellname[1:])
        return (number-1) * 8 + ('ABCDEFGH'.find(alphabet)+1) - 1 # 0부터 시작하니깐 -1

    def click_tab_resize(self, num_tab):
        """
        클릭한 탭이 Report 탭이면 프로그램 창의 크기를 키우고 다른 탭을 누르면 창 크기를 원상복구한다.

        Parameter
        ---------
        num_tab: int
            클릭한 탭의 번호(순서)
        """
        if num_tab==3:
            self.resize(1550, 907)
            self.tabWidget.resize(1535, 881)
            self.groupBox_8.resize(1505, 721)
        else:
            self.resize(1207, 907)
            self.tabWidget.resize(1191, 881)
            self.groupBox_8.resize(1161, 721)

    def click_btn_open_savelocation(self):
        """
        btn_open_savelocation의 클릭 이벤트. 해당 이벤트를 호출한 버튼의 이름에 할당된 저장 폴더를 연다.

        클릭된 버튼(sender())의 objectName()에 따라 할당된 저장폴더를 연다.
        """

        if self.sender().objectName() == 'btn_open_sheets' or self.sender().objectName() == 'btn_open_sheets_2' or self.sender().objectName() == 'btn_open_sheets_sub':
            dir = '/Sheets/'
        elif self.sender().objectName() == 'btn_open_RT' or self.sender().objectName() == 'btn_open_RT_2':
            dir = '/RT/'
        elif self.sender().objectName() == 'btn_opensavelocation_onsite' or self.sender().objectName() == 'btn_opensavelocation_storage':
            dir = '/ETC/'
        os.startfile(os.path.realpath(self.ddi_present.location_save+dir))

    def open_xls_file(self, filepath):
        if os.path.exists(filepath):
            self.showMinimized()
            self.dispatch_excel.Visible = True
            self.dispatch_excel.Workbooks.Open(filepath)
        else:
            QMessageBox.information(self, "Error", "File does not exist.")
    # info tab
    def set_line_texts(self, location_save="", analyst="", date=QtCore.QDate.currentDate().toString('yyyyMMdd')):
        """
        현재 작업의 저장위치, 담당자, 채취날짜를 인자로 받아 각각 해당하는 QLineEdit, QDateEdit 객체의 Text 속성에 할당한다.

        Parameters
        ------------
        location_save: str
            현재 작업의 저장 위치
        analyst: str
            현재 작업의 담당자
        date: QtCore.QDate
            현재 작업의 채취 날짜
        """

        self.line_savelocation.setText(location_save)
        self.line_analyst.setText(analyst)
        self.line_date.setDate(QtCore.QDate.fromString(date,'yyyyMMdd'))

    def update_info_table(self):
        """
        info_table에 ddi_present의 증거물 데이터프레임 값을 입력한다.
        """
        self.label_number_samples.setText('Number of samples : %d' % len(self.ddi_present.df_evidence))
        self.table_info.setColumnCount(10)
        self.table_info.setRowCount(len(self.ddi_present.df_evidence.index))
        self.table_info.setHorizontalHeaderLabels(['Index', '의뢰관서', '증거물 번호', '감정물', '분류'])
        idx_table = 0
        for (index, row) in self.ddi_present.df_evidence.iterrows():
            self.table_info.setItem(idx_table, 0, QTableWidgetItem(str(index)))
            self.table_info.setItem(idx_table, 1, QTableWidgetItem(str(row['의뢰관서'])))
            self.table_info.setItem(idx_table, 2, QTableWidgetItem(str(row['증거물번호'])))
            self.table_info.setItem(idx_table, 3, QTableWidgetItem(str(row['감정물'])))
            self.table_info.setItem(idx_table, 4, QTableWidgetItem(str(row['분류'])))
            idx_table = idx_table + 1

    def search_table(self, keyword, table):
        """
        해당 QTableWidget에서 keyword를 가진 아이템을 찾아 커서를 움직인다.

        Parameters
        ----------
        keyword : str
            검색할 키워드
        table : QTableWidget
             검색할 테이블 객체
        """

        result = table.findItems(keyword, QtCore.Qt.MatchContains)
        if result:
            table.setCurrentCell(result[0].row(), result[0].column())
        else:
            QMessageBox.information(self, "Notice", "No result.")

    def click_btn_search_info(self):
        self.search_table(self.line_search_info.text(), self.table_info)

    def click_btn_NFIS_login(self):
        """btn_NFIS_login의 클릭 이벤트. 외부 프로그램 NFIS_login을 실행한다."""

        self.run_external_app('NFIS_login')

    def click_btn_NFIS_revision_helper(self):
        """btn_NFIS_revision_helper의 클릭 이벤트. 외부 프로그램 NFIS_revision_helper를 실행한다."""

        self.run_external_app('NFIS_revision_helper')

    def click_btn_NFIS_tomato(self):
        """ btn_NFIS_tomato의 클릭 이벤트. 해당 프로젝트의 Tomato 엑셀 파일을 연다."""
        self.showMinimized()
        self.dispatch_excel.Visible = True
        self.dispatch_excel.Workbooks.Open(self.ddi_present.path_tomato)

    def click_btn_total_sheet(self):
        """btn_total_sheet의 클릭 이벤트. 해당 프로젝트의 total_sheet 엑셀 파일을 연다."""

        if os.path.exists(self.ddi_present.path_totalsheet):
            self.dispatch_excel.Visible = True
            self.dispatch_excel.Workbooks.Open(self.ddi_present.path_totalsheet)
        else:
            QMessageBox.information(self, "Error", "File does not exist.")

    # Sheets tab
    def load_samplesheets(self):
        """현재 ddi_present에 저장된 데이터에 따라 Sheet탭의 리스트와 콤보박스를 업데이트한다"""
        df_unassigned = self.ddi_present.df_evidence[self.ddi_present.df_evidence['분류'] == 'Unassigned']
        self.list_sample_all.clear()
        for (index, row) in df_unassigned.iterrows():
            self.list_sample_all.addItem(
                "{index:<10}{case:<15}{evidence}".format(index=index + 1, case=row['접수번호'], evidence=row['감정물']))
        self.combo_category.clear()
        for category in self.ddi_present.list_tag:
            self.combo_category.addItem(category)
        self.change_combo_category(self.ddi_present.list_tag[0])
        self.update_list_count(self.list_sample_all, self.label_count_sample_all)
        self.update_list_count(self.list_sample_partial, self.label_count_sample_partial)

    def generate_samplesheets(self, worksheet, df, filename, row_start, control = False, blank = False, ladder = False, macro=False, sheetname = ""):
        """
        해당 워크시트에 증거물 데이터프레임의 값을 입력한 후 지정된 파일 이름으로 저장한다.

        Parameters
        ----------
        worksheet: str
            데이터를 입력할 워크시트 파일의 경로
        df: DataFrame
            증거물 데이터프레임
        filename: str
            저장할 파일의 이름
        row_start: int
            샘플 시트 서식에서 값을 입력할 첫번째 행
        control: bool, optional
            CONTROL 행 추가 여부
        blank: bool, optional
            BLANK 행 추가 여부
        ladder: bool, optional
            LADDER 행 추가 여부
        """

        wb_form = load_workbook(worksheet, read_only=False, keep_vba=True)
        ws_form = wb_form.active if sheetname == "" else wb_form[sheetname]
        ws_form['C1'] = filename
        ws_form['H1'] = self.ddi_present.date
        ws_form['K1'] = self.ddi_present.analyst
        idx = 0
        for idx, line in df.iterrows():
            ws_form.cell(row=row_start + idx, column=2).value = line['의뢰관서']
            ws_form.cell(row=row_start + idx, column=3).value = line['증거물번호']
            ws_form.cell(row=row_start + idx, column=4).value = line['감정물'].split(':')[1]
            ws_form.cell(row=row_start + idx, column=5).value = line['분류']
        idx = idx + 1
        if control == True :
            ws_form.cell(row=row_start + idx, column=3).value = 'Control'
            idx = idx + 1
        if blank == True :
            ws_form.cell(row=row_start + idx, column=3).value = 'Blank'
            idx = idx + 1
        if ladder == True :
            ws_form.cell(row=row_start + idx, column=3).value = 'Ladder'
            idx = idx + 1
            ws_form.cell(row=row_start + idx, column=3).value = 'Ladder'
        ext = ".xlsm" if macro == True else ".xlsx"
        wb_form.save(
            self.ddi_present.location_save + '/Sheets/' + filename + ext)

    def update_list_count(self, target_qlistwidget, count_qlable):
        """
        list의 item 갯수를 qlabelwidget에 반영

        list의 내용이 변할 때 마다 해당 함수를 호출해서 해당 리스트와 연계된 qlabelwidget의 text 값을 변환.

        Parameters
        -----------
        taget_qlistwidget: qListWidget
            item의 걋수를 구할 list
        count_qlabel : qLabel
            item의 걋수를 반영할 label
        """

        count_qlable.setText(str(target_qlistwidget.count()))

    def change_combo_category(self, item):
        """
        combo_category의 변경 이벤트. combo_category에서 아이템을 선택했을 때 해당 분류에 속하는 증거물의 목록을 list_sample_partial에 반영한다.

        list_sample_partial을 초기화하고 선택된 item의 text에 해당하는 분류 값을 가진 증거물 데이터프레임의 값을 나열한다.

        Parameters
        -----------
        item: str
            해당 콤보박스에서 선택된 item
        """

        self.list_sample_partial.clear()
        for (index, row) in self.ddi_present.df_evidence[self.ddi_present.df_evidence['분류'] == item].iterrows():
            self.list_sample_partial.addItem(
                "{index:<10}{case:<15}{evidence}".format(index=index + 1, case=row['접수번호'], evidence=row['감정물']))
        self.update_list_count(self.list_sample_all, self.label_count_sample_all)
        self.update_list_count(self.list_sample_partial, self.label_count_sample_partial)

    def click_btn_import_modified_sample(self):
        """
        btn_import_modified_sample의 클릭 이벤트. 채취 후 수정한 NFIS파일을 읽고 ddi_present 객체에 저장한다. 그 후 증거물 목록을 list_sample_all에 반영한다.

        수정된 NFIS 파일의 경로를 입력받고 불러온 후 데이터프레임으로 전환하여 ddi_present 객체에 저장한다.
        증거물 데이터프레임을 초기화하고 첫번째 분류 값을 입력받는다.
        list_sample_all에 데이터프레임의 모든 열을 지정된 형식으로 추가한다.
        변경사항을 DataDNAIdentification.pickle 파일에 저장한다.
        """
        self.list_sample_all.clear()
        self.list_sample_partial.clear()
        self.combo_category.clear()

        self.line_import_raw_sample.setText(self.import_file(extension='xlsx(*.xlsx)', copy_needed=True))
        self.ddi_present.df_evidence = self.xls_to_dataframe(self.line_import_raw_sample.text())
        # 증거물 데이터 프레임 초기화. 작업에 필요한 행 추가.
        self.ddi_present.df_evidence['분류'] = 'Unassigned'
        self.ddi_present.df_evidence['증거물번호'] = self.ddi_present.df_evidence['접수번호'] + self.ddi_present.df_evidence['감정물'].apply(lambda x: '-'+x.split('증')[1].split('호')[0])
        self.sort_by_serial(self.ddi_present.df_evidence)
        self.ddi_present.nfis_loaded = True
        self.tabWidget.setTabEnabled(2, True)  # Resample tab 활성화
        self.tabWidget.setTabEnabled(3, True)   # Report tab 활성화
        self.tabWidget.setTabEnabled(4, True)   # Data tab 활성화
        self.combo_category.addItems(['LCN', 'MF', 'REF'])    # 기본 분류 설정
        for (index, row) in self.ddi_present.df_evidence.iterrows():
            self.list_sample_all.addItem("{index:<10}{case:<15}{evidence}".format(index=index+1, case=row['접수번호'], evidence=row['감정물']))
        # 감정서 DataFrame 생성
        self.ddi_present.df_report = self.ddi_present.df_evidence.copy()
        self.ddi_present.df_report['DB Type 1'] = ""
        self.ddi_present.df_report['DB Type 2'] = ""
        self.ddi_present.df_report['Y Type'] = ""
        self.ddi_present.df_report['DB_Hit'] = ""
        self.ddi_present.df_report['Matching Probability'] = ""
        self.ddi_present.df_report['Saliva'] = ""
        self.ddi_present.df_report['Semen'] = ""
        self.ddi_present.df_report['Blood'] = ""
        self.ddi_present.df_report['Return'] = ""
        self.ddi_present.df_report['Comment'] = ""
        self.load_resamplesheets()
        self.load_reportsheets()
        # 추가적으로 필요한 데이터는 이 칸에 추가
        # 저장 및 부가처리
        self.save() # ddi_present의 변경사항 저장

    def click_btn_add_category(self):
        """
        btn_add_category의 클릭 이벤트. 새로운 증거물 분류를 combo_category 추가한다.

        pyqt의 QInputDialog로 분류를 입력받고 중복 또는 공백 여부를 확인한다.
        문제가 없으면 ddi_present.list_tag와 combo_category에 해당 분류를 추가한다.
        문제가 있으면 -1을 반환

        Returns
        -------
        int
            입력에 문제가 있을 경우 -1을 반환
        """

        str_tag, ok = QInputDialog.getText(self, 'Tag Input Dialog', 'Enter a tag for the input file : ')
        if ok:
            if str_tag is "":
                QMessageBox.information(self, "Error", "No tag assigned.")
                return -1
            elif str_tag in self.ddi_present.list_tag:
                QMessageBox.information(self, "Error", "Already used tag.")
                return -1
            self.combo_category.addItem(str_tag)
            self.ddi_present.list_tag.append(str_tag)
        else:
            return -1
        self.combo_category.setCurrentIndex(self.combo_category.count()-1)
        self.change_combo_category(str_tag)
        self.update_list_count(self.list_sample_all, self.label_count_sample_all)
        self.update_list_count(self.list_sample_partial, self.label_count_sample_partial)

    def click_btn_remove_category(self):
        """
        btn_remove_category의 클릭 이벤트. 현재 선택된 증거물 분류를 combo_category에서 제거한다.
        """

        if self.combo_category.count()==1:  # 분류가 단 하나뿐이면
            QMessageBox.information(self, "Error", "At least one category is required.")
            return
        self.click_btn_remove_all()
        str_remove=self.combo_category.currentText()
        self.ddi_present.list_tag.remove(str_remove)
        self.combo_category.removeItem(self.combo_category.currentIndex())
        self.save()
        self.update_list_count(self.list_sample_all, self.label_count_sample_all)
        self.update_list_count(self.list_sample_partial, self.label_count_sample_partial)

    def click_btn_move_all(self):
        """
        btn_move_all의 클릭 이벤트. list_sample_all의 모든 내용을 list_sample_partial으로 옮긴다.

        list_sample_all의 모든 내용을 list_sample_partial으로 옮기고 리스트의 내용에 따라 증거물 데이터프레임의 분류 값을 갱신한다.
        """

        self.move_all_item(self.list_sample_all, self.list_sample_partial)
        self.update_df_sample(self.ddi_present.df_evidence, self.list_sample_partial, self.combo_category.currentText())
        self.update_list_count(self.list_sample_all, self.label_count_sample_all)
        self.update_list_count(self.list_sample_partial, self.label_count_sample_partial)

    def click_btn_remove_all(self):
        """
        btn_remove_all의 클릭 이벤트. list_sample_partial의 모든 내용을 list_sample_all으로 옮긴다.

        list_sample_partial의 모든 내용을 list_sample_all으로 옮기고 리스트의 내용에 따라 증거물 데이터프레임의 분류 값을 갱신한다.
        """

        self.move_all_item(self.list_sample_partial, self.list_sample_all)
        self.update_df_sample(self.ddi_present.df_evidence, self.list_sample_all, 'Unassigned')
        self.update_list_count(self.list_sample_all, self.label_count_sample_all)
        self.update_list_count(self.list_sample_partial, self.label_count_sample_partial)

    def click_btn_move(self):
        """
        btn_move의 클릭 이벤트. 선택된 아이템을 list_sample_all에서 list_sample_partial으로 옮긴다.

        선택된 아이템을 list_sample_all에서 list_sample_partial으로 옮기고 리스트의 내용에 따라 증거물 데이터프레임의 분류 값을 갱신한다.
        """

        self.move_items(self.list_sample_all, self.list_sample_partial, self.list_sample_all.selectedItems())
        self.update_df_sample(self.ddi_present.df_evidence, self.list_sample_partial, self.combo_category.currentText())
        self.update_list_count(self.list_sample_all, self.label_count_sample_all)
        self.update_list_count(self.list_sample_partial, self.label_count_sample_partial)

    def click_btn_remove(self):
        """
        btn_remove의 클릭 이벤트. 선택된 아이템을 list_sample_partial에서 list_sample_all으로 옮긴다.

        선택된 아이템을 list_sample_partial에서 list_sample_all으로 옮기고 리스트의 내용에 따라 증거물 데이터프레임의 분류 값을 갱신한다.
        """

        self.move_items(self.list_sample_partial, self.list_sample_all, self.list_sample_partial.selectedItems())
        self.update_df_sample(self.ddi_present.df_evidence, self.list_sample_all, 'Unassigned')
        self.update_list_count(self.list_sample_all, self.label_count_sample_all)
        self.update_list_count(self.list_sample_partial, self.label_count_sample_partial)

    # def click_btn_generate_samplesheets(self):
    #     """
    #     btn_generate_samplesheets의 클릭 이벤트. ddi_present의 증거물 데이터프레임에 저장된 데이터를 분류대로 나눠 샘플시트를 생성한다.
    #     """
    #
    #     groupby_tag = self.ddi_present.df_evidence.groupby('분류')
    #     for tag, group in groupby_tag:
    #         group = group.reset_index(drop=True)
    #         filename = self.ddi_present.date + '-' + self.ddi_present.analyst + '-' + tag
    #         self.generate_samplesheets(self.root + '/Form/form_samplesheet.xlsx', group, filename, 3, True, True, True)
    #     self.save()

    def click_btn_generate_totalsheet(self):
        """
        btn_generate_totalsheet의 클릭 이벤트. ddit_present의 증거물 데이터프레임에 저장된 데이터로 하나의 샘플시트를 생성한다
        """

        df_total = self.ddi_present.df_evidence[self.ddi_present.df_evidence['분류'] != 'Unassigned'] # 실험에 사용되지 않은 샘플을 제거한 데이터프레임 생성
        start_row = 3
        filename = self.ddi_present.date + '-' + self.ddi_present.analyst + '-' + 'TOTAL'
        self.generate_samplesheets(self.root + '/Form/form_sampletotalsheet.xlsm', pd.DataFrame({}), filename,
                                   start_row, False, False, False, True, "TOTAL")  # 우선 빈 시트를 생성

        groupby_tag = df_total.groupby('분류')
        for tag, group in groupby_tag:
            group = group.reset_index(drop=True)
            self.generate_samplesheets(self.ddi_present.location_save + '/Sheets/' + filename + ".xlsm", group, filename, start_row, False, False, False, True)
            start_row = start_row + len(group)
        self.save()
        self.open_xls_file(self.ddi_present.location_save + '/Sheets/' + filename + ".xlsm")
        QMessageBox.information(self, "Notice", "Work complete.")

    def click_btn_generate_RT_sheet_from_total(self):
        """
        btn_generate_RT_sheet_from_total 버튼의 클릭 이벤트. totalsheet 엑셀 파일의 TOTAL 시트에서 TYPE이 LCN, MF인 것만 추출하여 RT import 파일을 작성한다.
        """

        # RT 대상이 96개가 안되면 디폴트로 설정된 샘플링시트로 RT 시트를 생성. 넘으면 RT 시트를 생성할 샘플링 시트 경로를 입력 받음.
        count_LCN = self.ddi_present.df_evidence[self.ddi_present.df_evidence['분류'] == 'LCN'].shape[0] + self.ddi_present.df_evidence[self.ddi_present.df_evidence['분류'] == 'MF'].shape[0]
        if count_LCN < 96:
            path_samplingsheet = self.ddi_present.path_totalsheet
        else:
            path_samplingsheet = self.import_file(extension='xlsm(*.xlsm)', copy_needed=False, title="Choose a Sampling Sheet for RT")

        if not os.path.exists(path_samplingsheet):
            QMessageBox.information(self, "Error", "File does not exist.")
            return -1

        filename_RTsheet = self.ddi_present.location_save+'/RT/'+ path_samplingsheet.split('/')[-1].rstrip('.xlsm')+'_RT.txt'
        shutil.copyfile(self.root + '/Form/form_RT.txt', filename_RTsheet)
        with open(filename_RTsheet, mode='a') as f:
            wb_form = load_workbook(path_samplingsheet)
            ws_form = wb_form['TOTAL']
            for idx in range(96):
                wellname = self.idx_to_wellname(idx)    #idx는 0부터 시작
                str_serial = ws_form.cell(row=idx+3, column=3).value    # 샘플시트 상에서 3번째 열은 데이터의 시작열. 3번째 행은 증거물번호의 행
                str_type = ws_form.cell(row=idx+3, column=5).value
                if str_serial == None or str_type not in ['LCN', 'MF']: continue # 증거물 번호가 빈칸이면 다음 루프로 넘어감
                f.write(
                    '\t'.join(
                        [wellname, str_serial, '"RGB(255,153,204)"', 'T.IPC', '"RGB(255,0,0)"',
                         'UNKNOWN', 'JUN', 'QSY7']) + '\n')
                f.write(
                    '\t'.join([wellname, str_serial, '"RGB(255,153,204)"', 'T.Large Autosomal',
                               '"RGB(0,0,0)"',
                               'UNKNOWN', 'ABY', 'QSY7']) + '\n')
                f.write(
                    '\t'.join([wellname, str_serial, '"RGB(255,153,204)"', 'T.Small Autosomal',
                               '"RGB(0,128,0)"',
                               'UNKNOWN', 'VIC', 'NFQ-MGB']) + '\n')
                f.write(
                    '\t'.join(
                        [wellname, str_serial, '"RGB(255,153,204)"', 'T.Y', '"RGB(0,0,255)"',
                         'UNKNOWN', 'FAM', 'NFQ-MGB']) + '\n')
        QMessageBox.information(self, "Notice", "Work complete.")

    def click_btn_import_RT(self):
        """
        btn_Import_RT 버튼의 클릭 이벤트. RT 실험 결과 파일의 경로를 입력받는다. 그리고 해당 파일의 RT 실험 결과를 증거물 토탈샘플시트 파일에 복사한다.

        RT 결과 .xls 파일의 경로를 입력받고 해당 파일을 .xlsx파일로 전환한다. (openpyxl 라이브러리가 xlsx 파일만 지원)
        새로 생성된 xlsx파일의 경로를 line_import_RT에 반영한다.
        지정된 경로의 RT결과 엑셀 파일을 연다.
        샘플시트를 불러오고 RT 결과를 순차대로 입력한다.
        """

        # RT 대상이 96개가 안되면 디폴트로 설정된 샘플링시트로 RT 데이터를 복사. 넘으면 RT 데이터를 복사할 샘플링 시트 경로를 입력 받음.
        count_LCN = self.ddi_present.df_evidence[self.ddi_present.df_evidence['분류'] == 'LCN'].shape[0] + self.ddi_present.df_evidence[self.ddi_present.df_evidence['분류'] == 'MF'].shape[0]
        if count_LCN < 96:
            path_samplingsheet = self.ddi_present.path_totalsheet
        else:
            path_samplingsheet = self.import_file(extension='xlsm(*.xlsm)', copy_needed=False, title="Choose a Sampling Sheet for Import")

        if not os.path.exists(path_samplingsheet):
            QMessageBox.information(self, "Error", "File does not exist.")
            return -1

        filename = self.import_file(extension='xls(*.xls)', copy_needed=False, title="Choose a RT result file")
        if filename == None:
            QMessageBox.information(self, "Error", "Invalid file selection")
            return -1
        # xls 파일을 xlsx로 전환(for openpyxl)
        wb = self.dispatch_excel.Workbooks.Open(filename)
        wb.SaveAs(os.path.realpath(filename+'x'), FileFormat = 51) # 51 : xlsx 확장자
        wb.Close()
        # RT 결과값을 토탈샘플시트에 복사
        wb_data = load_workbook(os.path.realpath(filename+'x'))
        ws_data = wb_data.active
        blank = 0   # RT 결과에서 샘플명이 비어있는 칸을 세기 위한 카운터
        wb_total = load_workbook(path_samplingsheet, read_only=False, keep_vba=True)
        ws_total = wb_total["TOTAL"]
        for idx in range(96):
            idx_data = idx * 4 + 9 - 3 * blank
            serial = ws_data.cell(row=idx_data, column=2).value   # 샘플명, 1개의 샘플당 4개의 측정값이 존재. 9번째 줄부터 측정값 데이터가 시작됨. 샘플명이 없는 경우 공백 한 줄만 기록됨.
            wellname = ws_data.cell(row=idx_data, column=1).value
            if serial == "" or serial is None:    # 샘플명이 없는 경우 샘플명을 얻을 때 건너뛰어야 하는 열의 숫자를 증가하고 다음 샘플로 넘어간다.
                blank = blank + 1
            else:
                # idx + 3 : 3번째 열부터 데이터가 시작된다. idx는 0부터 카운팅\
                if wellname is not None:
                    idx_transformed = self.wellname_to_idx(wellname) + 3
                    ws_total.cell(row=idx_transformed, column=8).value = ws_data.cell(row=idx_data + 1, column=11).value if ws_data.cell(row=idx_data + 1, column=11).value !='' else 0  # Large autosomal
                    ws_total.cell(row=idx_transformed, column=6).value = ws_data.cell(row=idx_data + 2, column=11).value if ws_data.cell(row=idx_data + 2, column=11).value!='' else 0  # small autosomal
                    ws_total.cell(row=idx_transformed, column=7).value = ws_data.cell(row=idx_data + 3, column=11).value if ws_data.cell(row=idx_data + 3, column=11).value!='' else 0  # Y chromosome
        wb_total.save(path_samplingsheet)
        self.open_xls_file(path_samplingsheet)
        QMessageBox.information(self, "Notice", "Work complete.")

    def click_btn_auto_classification(self):
        """
        btn_btn_auto_classification의 클릭 이벤트. 키워드가 감정물명에 들어가 있으면 그 키워드에 해당하는 분류명을 자동으로 할당
        """
        dict_keyword = {'MF': ['F호', 'M호'],
                        'REF': ['혈액', '늑연골', '구강키트', '심낭혈'],
                        'Unassigned': ['소변', '슬라이드']}
        def search_keyword(evidence, dict_keyword):
            verdict = "LCN"
            for type, keywords in dict_keyword.items():
                for keyword in keywords:
                    if keyword in evidence:
                        verdict = type
            return verdict
        self.ddi_present.df_evidence['분류'] = self.ddi_present.df_evidence['감정물'].apply(lambda x:search_keyword(x,dict_keyword))
        self.load_samplesheets()
        self.save()

    # Resample tab
    def load_resamplesheets(self):
        self.list_resample_all.clear()
        for (index, row) in self.ddi_present.df_evidence.iterrows():
            self.list_resample_all.addItem(
                "{index:<10}{case:<15}{evidence}".format(index=index + 1, case=row['접수번호'], evidence=row['감정물']))
        self.update_list_count(self.list_resample_all, self.label_count_resample_all)
        self.update_list_count(self.list_resample_partial, self.label_count_resample_partial)

    def click_btn_move_all_resample(self):
        """
        btn_move_all_resample의 클릭 이벤트. list_resample_all의 모든 내용을 list_resample_partial으로 옮긴다.

        list_resample_all의 모든 내용을 list_resample_partial으로 옮긴다.
        """

        self.move_all_item(self.list_resample_all, self.list_resample_partial)
        self.update_list_count(self.list_resample_all, self.label_count_resample_all)
        self.update_list_count(self.list_resample_partial, self.label_count_resample_partial)

    def click_btn_remove_all_resample(self):
        """
        btn_remove_all_resample의 클릭 이벤트. list_resample_partial의 모든 내용을 list_resample_all으로 옮긴다.

        list_resample_partial의 모든 내용을 list_resample_all으로 옮긴다.
        """

        self.move_all_item(self.list_resample_partial, self.list_resample_all)
        self.update_list_count(self.list_resample_all, self.label_count_resample_all)
        self.update_list_count(self.list_resample_partial, self.label_count_resample_partial)

    def click_btn_move_resample(self):
        """
        btn_move_resample의 클릭 이벤트. 선택된 아이템을 list_resample_all에서 list_resample_partial으로 옮긴다.

        선택된 아이템을 list_resample_all에서 list_resample_partial으로 옮긴다.
        """

        self.move_items(self.list_resample_all, self.list_resample_partial, self.list_resample_all.selectedItems())
        self.update_list_count(self.list_resample_all, self.label_count_resample_all)
        self.update_list_count(self.list_resample_partial, self.label_count_resample_partial)

    def click_btn_remove_resample(self):
        """btn_remove_resample의 클릭 이벤트. 선택된 아이템을 list_resample_partial에서 list_resample_all으로 옮긴다."""

        self.move_items(self.list_resample_partial, self.list_resample_all, self.list_resample_partial.selectedItems())
        self.update_list_count(self.list_resample_all, self.label_count_resample_all)
        self.update_list_count(self.list_resample_partial, self.label_count_resample_partial)

    def click_btn_generate_resamplesheet(self):
        """
        btn_generate_resamplesheet의 클릭 이벤트. ddi_present의 증거물 데이터프레임에 저장된 데이터로  재실험시트를 생성한다
        """

        df_copy = self.ddi_present.df_evidence.copy()
        self.update_df_sample(df_copy, self.list_resample_partial, "RES")
        df_resample = df_copy[df_copy['분류']=="RES"].reset_index() # RESAMPLE한 증거물만 모아서 데이터프레임 생성

        start_row = 3
        filename = self.ddi_present.date + '-' + self.ddi_present.analyst + '-' + 'RESAMPLING'
        self.generate_samplesheets(self.root + '/Form/form_sampletotalsheet.xlsm', df_resample, filename,
                                   start_row, False, False, False, True, "TOTAL")
        QMessageBox.information(self, "Notice", "Work complete.")

    def click_btn_generate_RT_sheet_from_resamplesheet(self):
        """
        click_btn_generate_RT_sheet_from_resamplesheet 버튼의 클릭 이벤트. resamplesheet 엑셀 파일의 TOTAL 시트에서 RT import 파일을 작성한다.
        """

        if not os.path.exists(self.ddi_present.path_resamplesheet):
            QMessageBox.information(self, "Error", "File does not exist.")
            return -1
        filename_RTsheet = self.ddi_present.location_save+'/RT/'+ self.ddi_present.path_resamplesheet.split('/')[-1].rstrip('.xlsm')+'_RT.txt'
        shutil.copyfile(self.root + '/Form/form_RT.txt', filename_RTsheet)
        with open(filename_RTsheet, mode='a') as f:
            wb_form = load_workbook(self.ddi_present.path_resamplesheet)
            ws_form = wb_form['TOTAL']
            for idx in range(96):
                wellname = self.idx_to_wellname(idx)    #idx는 0부터 시작
                str_serial = ws_form.cell(row=idx+3, column=3).value    # 샘플시트 상에서 3번째 열은 데이터의 시작열. 3번째 행은 증거물번호의 행
                str_type = ws_form.cell(row=idx+3, column=5).value
                if str_serial == None or str_type not in ['RES']: continue # 증거물 번호가 빈칸이면 다음 루프로 넘어감
                f.write(
                    '\t'.join(
                        [wellname, str_serial, '"RGB(255,153,204)"', 'T.IPC', '"RGB(255,0,0)"',
                         'UNKNOWN', 'JUN', 'QSY7']) + '\n')
                f.write(
                    '\t'.join([wellname, str_serial, '"RGB(255,153,204)"', 'T.Large Autosomal',
                               '"RGB(0,0,0)"',
                               'UNKNOWN', 'ABY', 'QSY7']) + '\n')
                f.write(
                    '\t'.join([wellname, str_serial, '"RGB(255,153,204)"', 'T.Small Autosomal',
                               '"RGB(0,128,0)"',
                               'UNKNOWN', 'VIC', 'NFQ-MGB']) + '\n')
                f.write(
                    '\t'.join(
                        [wellname, str_serial, '"RGB(255,153,204)"', 'T.Y', '"RGB(0,0,255)"',
                         'UNKNOWN', 'FAM', 'NFQ-MGB']) + '\n')
        QMessageBox.information(self, "Notice", "Work complete.")

    def click_btn_import_RT_resample(self):
        """
        click_btn_import_RT_resample 버튼의 클릭 이벤트. RT 실험 결과 파일의 경로를 입력받는다. 그리고 해당 파일의 RT 실험 결과를 증거물 토탈샘플시트 파일에 복사한다.

        RT 결과 .xls 파일의 경로를 입력받고 해당 파일을 .xlsx파일로 전환한다. (openpyxl 라이브러리가 xlsx 파일만 지원)
        새로 생성된 xlsx파일의 경로를 line_import_RT에 반영한다.
        지정된 경로의 RT결과 엑셀 파일을 연다.
        샘플시트를 불러오고 RT 결과를 순차대로 입력한다.
        """

        filename = self.import_file(copy_needed=False)
        if filename == None:
            QMessageBox.information(self, "Error", "Invalid file selection")
            return
        # xls 파일을 xlsx로 전환(for openpyxl)
        wb = self.dispatch_excel.Workbooks.Open(filename)
        wb.SaveAs(os.path.realpath(filename+'x'), FileFormat = 51) # 51 : xlsx 확장자
        wb.Close()
        # RT 결과값을 토탈샘플시트에 복사
        wb_data = load_workbook(os.path.realpath(filename+'x'))
        ws_data = wb_data.active
        blank = 0   # RT 결과에서 샘플명이 비어있는 칸을 세기 위한 카운터
        wb_total = load_workbook(self.ddi_present.path_resamplesheet, read_only=False, keep_vba=True)
        ws_total = wb_total["TOTAL"]
        for idx in range(96):
            idx_data = idx * 4 + 9 - 3 * blank
            serial = ws_data.cell(row=idx_data, column=2).value   # 샘플명, 1개의 샘플당 4개의 측정값이 존재. 9번째 줄부터 측정값 데이터가 시작됨. 샘플명이 없는 경우 공백 한 줄만 기록됨.
            wellname = ws_data.cell(row=idx_data, column=1).value
            if serial == "" or serial is None:    # 샘플명이 없는 경우 샘플명을 얻을 때 건너뛰어야 하는 열의 숫자를 증가하고 다음 샘플로 넘어간다.
                blank = blank + 1
            else:
                # idx + 3 : 3번째 열부터 데이터가 시작된다. idx는 0부터 카운팅
                if wellname is not None:
                    idx_transformed = self.wellname_to_idx(wellname) + 3
                    ws_total.cell(row=idx_transformed, column=8).value = ws_data.cell(row=idx_data + 1, column=11).value if ws_data.cell(row=idx_data + 1, column=11).value !='' else 0  # Large autosomal
                    ws_total.cell(row=idx_transformed, column=6).value = ws_data.cell(row=idx_data + 2, column=11).value if ws_data.cell(row=idx_data + 2, column=11).value!='' else 0  # small autosomal
                    ws_total.cell(row=idx_transformed, column=7).value = ws_data.cell(row=idx_data + 3, column=11).value if ws_data.cell(row=idx_data + 3, column=11).value!='' else 0  # Y chromosome
        wb_total.save(self.ddi_present.path_resamplesheet)
        QMessageBox.information(self, "Notice", "Work complete.")

    # Report tab
    def load_reportsheets(self):
        """감정서 데이터프레임 내의 접수번호를 리스트로 만들고 combo_report_cases에 반영한다"""

        list_cases = list(self.ddi_present.df_report['접수번호'].unique())
        self.combo_report_cases.addItems(list_cases)
        self.change_combo_report_cases(list_cases[0])

    def change_combo_report_cases(self, item):
        """
        선택된 사건번호에 따라 감정서 테이블을 갱신한다.

        Parameters
        ---------
        item: str
            해당 콤보박스에서 선택된 item의 text
        """

        self.table_report.clear()
        self.update_table_report(item)
        self.label_picture.clear
        self.load_list_images(item)

    def click_btn_report_next(self):
        """combo_report_cases의 다음 item 선택"""

        next_row = self.combo_report_cases.currentIndex() + 1 if self.combo_report_cases.currentIndex()+1 is not self.combo_report_cases.count() else 0     # 마지막 아이템일 때 처음으로
        self.combo_report_cases.setCurrentIndex(next_row)
        self.change_combo_report_cases(self.combo_report_cases.currentText())

    def click_btn_load_tomato(self):
        """Tomato 엑셀 파일의 combined_result 탭에서 DNA profile 데이터를 NFS_DNA 클래스 상에 불러온다. Y23 Tomamto 파일이 있다면 해당 파일의 데이터도 불러온다."""
        self.ddi_present.combined_result = NFS_DNA.CombinedResult()
        self.ddi_present.combined_result.load_tomato(self.ddi_present.path_tomato)
        for sample_name in self.ddi_present.combined_result.info.index: # 불러온 데이터를 Report 테이블에 반영
            print(sample_name)
            idx_match = self.ddi_present.df_report[self.ddi_present.df_report['증거물번호'] == sample_name].index
            if len(idx_match)==0:
                continue
            else:
                idx_match = idx_match[0]
            print(idx_match)
            self.ddi_present.df_report.loc[idx_match, 'DB Type 1'] = self.ddi_present.combined_result.info.loc[sample_name, 'DB Type 1']
            self.ddi_present.df_report.loc[idx_match, 'DB Type 2'] = self.ddi_present.combined_result.info.loc[sample_name, 'DB Type 2']
            self.ddi_present.df_report.loc[idx_match, 'Matching Probability'] = self.ddi_present.combined_result.info.loc[sample_name, 'Matching Probability']
        self.change_combo_report_cases(self.combo_report_cases.currentText())
        if os.path.isfile(self.ddi_present.path_tomato_y23):
            self.ddi_present.combined_result_y23.load_tomato(self.ddi_present.path_tomato_y23)
        QMessageBox.information(self, "Notice", "Work complete.")

    def update_table_report(self, number_case):
        """table_report에 ddi_present의 df_report값을 입력한다."""

        df_case = self.ddi_present.df_report[self.ddi_present.df_report['접수번호'] == number_case]
        self.table_report.setColumnCount(13)
        self.table_report.setRowCount(len(df_case))
        self.table_report.setHorizontalHeaderLabels(['index','증거물번호', '감정물', 'DB Type 1' , 'DB Type 2', 'Y Type', 'Matching Probability', 'Saliva', 'Semen', 'Blood', 'DB_Hit', 'Return', 'Comment'])
        idx_table = 0
        for (index, row) in df_case.iterrows():
            self.table_report.setItem(idx_table, 0, QTableWidgetItem(str(index)))
            self.table_report.setItem(idx_table, 1, QTableWidgetItem(str(row['증거물번호'])))
            self.table_report.setItem(idx_table, 2, QTableWidgetItem(str(row['감정물'])))
            self.table_report.setItem(idx_table, 3, QTableWidgetItem(str(row['DB Type 1'])))
            self.table_report.setItem(idx_table, 4, QTableWidgetItem(str(row['DB Type 2'])))
            self.table_report.setItem(idx_table, 5, QTableWidgetItem(str(row['Y Type'])))
            self.table_report.setItem(idx_table, 6, QTableWidgetItem(str(row['Matching Probability'])))
            self.table_report.setItem(idx_table, 7, QTableWidgetItem(str(row['Saliva'])))
            self.table_report.setItem(idx_table, 8, QTableWidgetItem(str(row['Semen'])))
            self.table_report.setItem(idx_table, 9, QTableWidgetItem(str(row['Blood'])))
            self.table_report.setItem(idx_table, 10, QTableWidgetItem(str(row['DB_Hit'])))
            self.table_report.setItem(idx_table, 11, QTableWidgetItem(str(row['Return'])))
            self.table_report.setItem(idx_table, 12, QTableWidgetItem(str(row['Comment'])))
            idx_table = idx_table + 1
        self.table_report.setColumnHidden(0, True)

    def cellchange_table_report(self, row, col):
        """증거물 테이블의 내용이 변경되면 변경된 내용을 감정서 데이터프레임에 반영한다 """

        idx = int(self.table_report.item(row ,0).text())    # dataframe의 index
        col_name = self.table_report.horizontalHeaderItem(col).text()
        text = self.table_report.item(row, col).text()
        self.ddi_present.df_report.loc[idx, col_name] = text

    def load_image(self, path):
        """입력받은 경로의 이미지를 label 객체에 띄운다"""
        pic = QPixmap()
        pic.load(path)
        pic = pic.scaled(self.label_picture.width(), self.label_picture.height())
        self.label_picture.setPixmap(pic)

    def load_list_images(self, case_number):
        """해당 사건번호을 파일이름에 포함하는 이미지의 이름을 찾아서 list_images에 반영한다"""
        self.list_picture.clear()
        list_img = [filename for filename in os.listdir(self.ddi_present.path_picture) if case_number in filename]
        for filename in list_img:
            item = QListWidgetItem()
            item.setText(filename.split('.')[0]) #QtGui.QApplication.translate("Dialog", x, None, QtGui.QApplication.UnicodeUTF8)
            item.setFlags(item.flags() | QtCore.Qt.ItemIsUserCheckable)
            item.setCheckState(QtCore.Qt.Unchecked)
            self.list_picture.addItem(item)
        if len(list_img)!=0:
            self.load_image(self.ddi_present.path_picture+'\\'+self.list_picture.item(0).text())
            self.list_picture.item(0).setSelected(True)

    def click_list_picture_item(self, item):
        """list_images에서 클릭된 아이템을 파일이름으로 가지는 이미지를 label 객체에 띄운다"""
        self.load_image(self.ddi_present.path_picture + '\\' + item.text())

    def click_btn_generate_report(self):
        """
        생성할 감정서 종류와 선택된 사건번호의 데이터를 토대로 해당 감정서를 작성한다.
        """

        number_case = self.combo_report_cases.currentText()
        type_report = self.combo_report_type.currentText()
        self.generate_report(number_case, type_report)
        QMessageBox.information(self, "보고서 생성", "생성 완료")

    def generate_report(self, num_case, type_report):
        """
        선택된 사건번호를 생성할 감정서 종류에 맞춰 감정서 hwp 파일을 생성한다.

        Parameters
        ----------
        num_case: str
            사건번호
        type_report: str
            감정서 종류

        Nested Functions
        ----------------
        generate_file(hwp_control, path_form, filename, y23=False):
            선택된 감정서 종류의 폼을 입력받은 이름의 파일로 생성하고 list_images에서 선택된 image를 삽입.
        load_profile(df_profile):
            입력받은  df의 증거물번호에 해당하는 프로파일과 프로파일의 기타사항을 반환
        load_profile_y23(df_profile):
            입력받은  df의 증거물번호에 해당하는 y23 프로파일과 프로파일의 기타사항을 반환
        write_alleles(hwp_control, num_slot, name_col, profile, list_marker, y23=False)
            좌위 테이블의 지정된 slot에 프로파일을 입력한다.
        link_num_evidence(df_target):
            입력받은 df의 증거물 번호를 추출해서 감정서에서 쓸 형태로 변환한다.
        gender_to_text(profile):
            profile의 Amelogenin 값에 따라 여성 혹은 남성 text를 반환한다.
        """

        def generate_file(hwp_control, path_form, filename, y23=False):
            shutil.copyfile(path_form, filename)
            hwp_control.Open(filename, "HWP", None)
            list_img_checked = []
            for row_number in range(self.list_picture.count()):
                if self.list_picture.item(row_number).checkState() == QtCore.Qt.Checked:
                    list_img_checked.append(self.list_picture.item(row_number).text())
            # 그림 테이블로 이동
            hwp_control.Run("MoveDocBegin")
            hwp_control.Run('MovePageDown')
            hwp_control.Run('MovePageDown')
            if y23==True:
                hwp_control.Run('MovePageDown')
            hwp_control.Run("MoveDown")
            # 체크된 사진을 사진 테이블로 복사
            for i, filename_img in enumerate(list_img_checked):
                filepath = rf"{self.ddi_present.path_picture}{filename_img}.jpg"
                hwp_control.InsertPicture(filepath, Embedded=True, sizeoption=3)
                # hwp.Run('TableRightCellAppend')
                sleep(0.1)
                if i % 2 == 0:
                    hwp_control.Run("TableAppendRow")
                else:
                    hwp_control.Run("MoveDown")
                num_extracted = '-'.join(filename_img.split('-')[3:])
                num_extracted = num_extracted.split('+')
                num_extracted = [re.sub('\d+', '증\g<0>호', x).replace('-', '~') for x in num_extracted]
                num_extracted = ', '.join(num_extracted)
                hwp_control.HAction.GetDefault("InsertText", hwp_control.HParameterSet.HInsertText.HSet)
                hwp_control.HParameterSet.HInsertText.Text = num_extracted
                hwp_control.HAction.Execute("InsertText", hwp_control.HParameterSet.HInsertText.HSet)
                hwp_control.Run("TableRightCellAppend")
                if i % 2 == 0:
                    hwp_control.Run("MoveUp")
                else:
                    pass
            hwp_control.Run("MoveDocBegin")

        # def load_profile_y23(df_profile):
        #     try:
        #         num_evidence =  df_profile.reset_index().loc[0, '증거물번호']
        #         profile_ref, str_etc_y23 = self.ddi_present.combined_result_y23.profiles[
        #             num_evidence].transform_to_str(False)
        #         return profile_ref, str_etc_y23
        #     except KeyError:
        #         QMessageBox.information(self, "Error", "No Profile Data.")
        #         hwp_control.Quit()
        #         raise KeyError

        def write_alleles(info, hwp_control, num_slot, list_marker, nickname="", y23=False):
            name_col = '{0}\r({1})'.format(info['link_num_evidence'].replace(" 및 ", ", "), nickname) if nickname!="" else '{0}'.format(info['link_num_evidence'].replace(" 및 ", ", "))
            profile = info['profile']

            # 페이지 쪽으로 나누고 PageDown으로 다음다음 넘어가면서..
            # Locus 테이블의 좌위 입력 부위로 이동
            hwp_control.Run("MoveDocBegin")
            hwp_control.MovePos(2) #캐럿을 문서 처음으로 이동
            hwp_control.Run('MovePageDown')
            if y23==True:
                hwp_control.Run('MovePageDown')
            hwp_control.Run("MoveDown")
            hwp_control.Run("MoveDown")
            hwp_control.Run("MoveRight")
            hwp_control.Run("MoveDown")
            for i in range(num_slot):
                hwp_control.MovePos(101) #캐럿을 오른쪽 셀로 이동
            hwp_control.HAction.GetDefault("InsertText", hwp_control.HParameterSet.HInsertText.HSet)
            hwp_control.HParameterSet.HInsertText.Text = name_col
            hwp_control.HAction.Execute("InsertText", hwp_control.HParameterSet.HInsertText.HSet)
            hwp_control.MovePos(103)  #캐럿을 아래쪽 셀로 이동
            for loci in list_marker:
                hwp_control.HAction.GetDefault("InsertText", hwp_control.HParameterSet.HInsertText.HSet)
                hwp_control.HParameterSet.HInsertText.Text = profile[loci]
                hwp_control.HAction.Execute("InsertText", hwp_control.HParameterSet.HInsertText.HSet)
                hwp_control.MovePos(103)  #캐럿을 아래쪽 셀로 이동
            hwp_control.Run("MoveDocBegin")

        def gender_to_text(profile):
            if profile['AMEL'] == 'XX':
                return '여성'
            else:
                return '남성'

        def process_info(df_case = pd.DataFrame({}),
                         type="", is_REF=False, REF_type=None, is_y =False):

            combined_result =  self.ddi_present.combined_result if is_y==False else self.ddi_present.combined_result_y23
            info_combined_result = combined_result.info
            list_marker = combined_result.list_marker_ordered
            type_colname = 'DB Type 1' if is_y==False else 'Y Type'

            def link_num_evidence(df_target):
                """증거물 번호의 리스트를 감정서에 넣을 포멧으로 변환하여 반환한다."""

                p = re.compile('증(.+)호')
                list_serial = df_target['감정물'].apply(lambda x: str(p.search(x).group(1))).tolist()
                list_idx = list(df_target.index)
                stack_serial = []
                stack_idx = []
                list_result = []
                if len(list_serial) == 1:
                    list_result.append('증{0}호'.format(list_serial[0]))
                elif len(list_serial) == 2:
                    list_result.append('증{0}호 및 증{1}호'.format(list_serial[0], list_serial[1]))
                else:
                    stack_serial.append(list_serial[0])
                    stack_idx.append(list_idx[0])
                    for idx, target in enumerate(list_serial[1:]):
                        if (stack_idx[-1] + 1) == list_idx[(idx + 1)]:
                            stack_idx.append(list_idx[(idx + 1)])
                            stack_serial.append(target)
                        else:
                            if len(stack_serial) == 1:
                                list_result.append('증{0}호'.format(stack_serial[0]))
                                stack_serial.clear()
                                stack_idx.clear()
                                stack_serial.append(target)
                                stack_idx.append(list_idx[(idx + 1)])

                            elif len(stack_serial) == 2:
                                list_result.append('증{0}호'.format(stack_serial[0]))
                                list_result.append('증{0}호'.format(stack_serial[1]))
                                stack_serial.clear()
                                stack_idx.clear()
                                stack_serial.append(target)
                                stack_idx.append(list_idx[(idx + 1)])
                            else:
                                list_result.append('증{0}호~증{1}호'.format(stack_serial[0], stack_serial[-1]))
                                stack_serial.clear()
                                stack_idx.clear()
                                stack_serial.append(target)
                                stack_idx.append(list_idx[(idx + 1)])
                    if len(stack_serial) != 0:
                        if len(stack_serial) == 1:
                            list_result.append('증{0}호'.format(stack_serial[0]))
                            stack_idx.clear()
                            stack_serial.clear()
                        elif len(stack_serial) == 2:
                            list_result.append('증{0}호'.format(stack_serial[0]))
                            list_result.append('증{0}호'.format(stack_serial[1]))
                            stack_idx.clear()
                            stack_serial.clear()
                        else:
                            list_result.append('증{0}호~증{1}호'.format(stack_serial[0], stack_serial[-1]))
                            stack_idx.clear()
                            stack_serial.clear()
                return ', '.join(list_result)

            def load_profile(df_profile, is_y = False):
                try:
                    num_evidence = df_profile.reset_index().loc[0, '증거물번호']
                    if is_y == False:
                        profile, str_etc = combined_result.profiles[num_evidence].transform_to_str(True)
                        profile['AMEL'] = profile['AMEL'].replace('-', '')
                    else:
                        profile, str_etc = combined_result.profiles[num_evidence].transform_to_str(False)
                    return profile, str_etc
                except KeyError:
                    QMessageBox.information(self, "Error", "No Profile Data.")
                    raise KeyError

            info = {}
            profile = {}
            str_etc = ""
            mp = None

            df_target = df_case[df_case[type_colname]==type]
            link_num_evidence = link_num_evidence(df_target)
            if type == 'ND':
                for loci in list_marker:
                    profile[loci] = "ND"
            elif is_REF==True:
                profile, str_etc = load_profile(df_target, is_y=is_y)
                num_evidence = df_target.reset_index().loc[0, '증거물번호']
                mp = str(info_combined_result.loc[num_evidence, 'Matching Probability']) if is_y==False else ""
            elif is_REF==False:
                df_ref = df_case[df_case[type_colname]==REF_type]
                profile, str_etc = load_profile(df_ref, is_y=is_y)

            info['link_num_evidence'] = link_num_evidence
            info['profile'] = profile
            info['str_etc'] = str_etc
            info['mp'] = mp
            return info

        df_case = self.ddi_present.df_report[self.ddi_present.df_report['접수번호'] == num_case].reset_index(drop=True)
        hwp_control = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
        hwp_control.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")  # 보안 모듈 적용(파일 열고 닫을 때 팝업이 안나타나게)
        filename_new = self.ddi_present.location_save + '/Reports/' + num_case + ".hwp"
        if type_report=='ND':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=False)
            info_ND = process_info(df_case, type='ND')
            hwp_control.PutFieldText("num_evidence_result{{0}}", info_ND['link_num_evidence'])
            hwp_control.PutFieldText("num_evidence_locus{{0}}", info_ND['link_num_evidence'].replace(" 및 ", ", "))
        elif type_report=='부검':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=False)
            info_D = process_info(df_case,
                                  type='D', is_REF=True)
            write_alleles(info=info_D, hwp_control=hwp_control, num_slot=1,
                          list_marker= self.ddi_present.combined_result.list_marker_ordered,
                          nickname="변사자",
                          y23=False)
            hwp_control.PutFieldText("num_evidence_result{{0}}", info_D['link_num_evidence'])
            hwp_control.PutFieldText("gender_result{{0}}", gender_to_text(info_D['profile']))
            hwp_control.PutFieldText("text_etc_locus{{0}}", info_D['str_etc'])
        elif type_report=='피해자 일치':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=False)

            info_V = process_info(df_case,
                                  type='V', is_REF=True)    # 피해자 대조
            info_v = process_info(df_case,
                                  type='v', is_REF=False, REF_type='V') #피해자 일치
            write_alleles(info=info_v, hwp_control=hwp_control, num_slot=1,
                          list_marker= self.ddi_present.combined_result.list_marker_ordered[:-3]) #부검이 아닌 REF는 SE33, PENTA_D, PENTA_E 제외
            write_alleles(info=info_V, hwp_control=hwp_control, num_slot=2,
                          list_marker= self.ddi_present.combined_result.list_marker_ordered[:-3],
                          nickname='피해자')
            hwp_control.PutFieldText("num_evidence_result{{0}}", info_v['link_num_evidence'])
            hwp_control.PutFieldText("gender_result{{0}}", gender_to_text(info_V['profile']))
            hwp_control.PutFieldText("float_mp_report{{0}}", info_V['mp'][:3]) # Matching Probalbility의 소수 부분
            hwp_control.PutFieldText("exp_mp_report{{0}}", info_V['mp'].split('+')[1]) #Matching Probability의 지수 부분
            hwp_control.PutFieldText("text_etc_locus{{0}}", info_V['str_etc'])
        elif type_report == 'ND w/ 피해자 일치':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=False)

            info_V = process_info(df_case,
                                  type='V', is_REF=True)    # 피해자 대조
            info_v = process_info(df_case,
                                  type='v', is_REF=False, REF_type='V') #피해자 일치
            info_ND = process_info(df_case, type='ND')
            write_alleles(info=info_v, hwp_control=hwp_control, num_slot=1,
                          list_marker= self.ddi_present.combined_result.list_marker_ordered[:-3]) #부검이 아닌 REF는 SE33, PENTA_D, PENTA_E 제외
            write_alleles(info=info_ND, hwp_control=hwp_control, num_slot=2,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])
            write_alleles(info=info_V, hwp_control=hwp_control, num_slot=3,
                          list_marker= self.ddi_present.combined_result.list_marker_ordered[:-3],
                          nickname='피해자')
            hwp_control.PutFieldText("num_evidence_v_result{{0}}", info_v['link_num_evidence'])
            hwp_control.PutFieldText("num_evidence_ND_result{{0}}", info_ND['link_num_evidence'])
            hwp_control.PutFieldText("gender_result{{0}}", gender_to_text(info_V['profile']))
            hwp_control.PutFieldText("float_mp_report{{0}}", info_V['mp'][:3])  # Matching Probalbility의 소수 부분
            hwp_control.PutFieldText("exp_mp_report{{0}}", info_V['mp'].split('+')[1])  # Matching Probability의 지수 부분
            hwp_control.PutFieldText("text_etc_locus{{0}}", 'ND : 디엔에이형이 검출되지 않음.\r' + info_V['str_etc'])
        elif type_report == 'ND w/ 피해자 불일치':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=False)
            info_V = process_info(df_case,
                                  type='V', is_REF=True)  # 피해자 대조
            info_ND = process_info(df_case, type='ND')
            write_alleles(info=info_ND, hwp_control=hwp_control, num_slot=1,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])
            write_alleles(info=info_V, hwp_control=hwp_control, num_slot=2,
                          list_marker= self.ddi_present.combined_result.list_marker_ordered[:-3],
                          nickname='피해자')
            hwp_control.PutFieldText("num_evidence_ND_result{{0}}", info_ND['link_num_evidence'])
            hwp_control.PutFieldText("text_etc_locus{{0}}", 'ND : 디엔에이형이 검출되지 않음.\r' + info_V['str_etc'])
        elif type_report=='혼합형':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=True)
            info_MX = process_info(df_case,
                                   type='MX', is_REF=True)
            write_alleles(info=info_MX, hwp_control=hwp_control, num_slot=1,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])  # 부검이 아닌 REF는 SE33, PENTA_D, PENTA_E 제외
            df_Y = df_case[df_case['Y Type'] == 'MX']
            if (len(df_Y) != 0):
                info_Y = process_info(df_case, type='MX', is_REF=True, is_y=True)
                write_alleles(info=info_Y, hwp_control=hwp_control, num_slot=1,
                              list_marker=self.ddi_present.combined_result_y23.list_marker_ordered,
                              y23=True)  # 부검이 아닌 REF는 SE33, PENTA_D, PENTA_E 제외
                hwp_control.PutFieldText("text_etc_locus_y23{{0}}", info_Y['str_etc'])
            hwp_control.PutFieldText("text_etc_locus{{0}}", info_MX['str_etc'])
        elif type_report=='피의자 일치' or type_report=='피의자 불일치':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=True)
            info_S = process_info(df_case, type='S', is_REF=True)
            write_alleles(info=info_S, hwp_control=hwp_control, num_slot=1,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3],
                          nickname='피의자')
            df_Y = df_case[df_case['Y Type'] == 'S']
            if (len(df_Y) != 0):
                info_Y = process_info(df_case, type='S', is_REF=True, is_y=True)
                write_alleles(info=info_Y, hwp_control=hwp_control, num_slot=1,
                              list_marker=self.ddi_present.combined_result_y23.list_marker_ordered,
                              nickname='피의자',
                              y23=True)  # 부검이 아닌 REF는 SE33, PENTA_D, PENTA_E 제외
                hwp_control.PutFieldText("text_etc_locus_y23{{0}}", info_Y['str_etc'])
            if type_report=='피의자 일치':
                hwp_control.PutFieldText("float_mp_report{{0}}", info_S['mp'][:3])  # Matching Probalbility의 소수 부분
                hwp_control.PutFieldText("exp_mp_report{{0}}", info_S['mp'].split('+')[1])  # Matching Probability의 지수 부분
            hwp_control.PutFieldText("text_etc_locus{{0}}", info_S['str_etc'])
        elif type_report=='친자관계 일치':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=True)
            info_R = process_info(df_case, type='R', is_REF=True)
            write_alleles(info=info_R, hwp_control=hwp_control, num_slot=1,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered,
                          nickname='관계자')
            df_Y = df_case[df_case['Y Type'] == 'R']
            if (len(df_Y) != 0):
                info_Y = process_info(df_case, type='R', is_REF=True, is_y=True)
                write_alleles(info=info_Y, hwp_control=hwp_control, num_slot=1,
                              list_marker=self.ddi_present.combined_result_y23.list_marker_ordered,
                              nickname='관계자',
                              y23=True)  # 부검이 아닌 REF는 SE33, PENTA_D, PENTA_E 제외
                hwp_control.PutFieldText("text_etc_locus_y23{{0}}", info_Y['str_etc'])
            hwp_control.PutFieldText("text_etc_locus{{0}}", info_R['str_etc'])
        elif type_report=='친자관계 일치(부검)':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=True)
            info_D = process_info(df_case, type='D', is_REF=True)
            info_R = process_info(df_case, type='R', is_REF=True)
            str_etc = info_D['str_etc']+'\r'+info_R['str_etc']
            write_alleles(info=info_R, hwp_control=hwp_control, num_slot=1,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered,
                          nickname='관계자')
            write_alleles(info=info_D, hwp_control=hwp_control, num_slot=2,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered,
                          nickname='변사자')
            df_Y = df_case[df_case['Y Type'] == 'R']
            if (len(df_Y) != 0):
                info_Y = process_info(df_case, type='R', is_REF=True, is_y=True)
                write_alleles(info=info_Y, hwp_control=hwp_control, num_slot=1,
                              list_marker=self.ddi_present.combined_result_y23.list_marker_ordered,
                              nickname='관계자',
                              y23=True)  # 부검이 아닌 REF는 SE33, PENTA_D, PENTA_E 제외
                hwp_control.PutFieldText("text_etc_locus_y23{{0}}", info_Y['str_etc'])
            hwp_control.PutFieldText("num_evidence_result{{0}}", info_D['link_num_evidence'])
            hwp_control.PutFieldText("gender_result{{0}}", gender_to_text(info_D['profile']))
            hwp_control.PutFieldText("text_etc_locus{{0}}", str_etc)
        elif type_report=='C 검출(검색결과 X)':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=False)
            info_C = process_info(df_case, type='C', is_REF=True)
            write_alleles(info=info_C, hwp_control=hwp_control, num_slot=1,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])
            hwp_control.PutFieldText("num_evidence_result{{0}}", info_C['link_num_evidence'])
            hwp_control.PutFieldText("gender_result{{0}}", gender_to_text(info_C['profile']))
            hwp_control.PutFieldText("text_etc_locus{{0}}", info_C['str_etc'])
        elif type_report == 'C 검출 w/ ND(검색결과 X)':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=False)
            info_C = process_info(df_case, type='C', is_REF=True)
            info_ND = process_info(df_case, type='ND')
            write_alleles(info=info_C, hwp_control=hwp_control, num_slot=1,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])
            write_alleles(info=info_ND, hwp_control=hwp_control, num_slot=2,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])
            hwp_control.PutFieldText("num_evidence_result{{0}}", info_C['link_num_evidence'])
            hwp_control.PutFieldText("num_evidence_ND_result{{0}}", info_ND['link_num_evidence'])
            hwp_control.PutFieldText("gender_result{{0}}", gender_to_text(info_C['profile']))
            hwp_control.PutFieldText("text_etc_locus{{0}}", info_C['str_etc'])
            hwp_control.PutFieldText("num_evidence_ND_result{{0}}", info_ND['link_num_evidence'])
            hwp_control.PutFieldText("text_etc_locus{{0}}", 'ND : 디엔에이형이 검출되지 않음.\r' + info_C['str_etc'])
        elif type_report=='C 검출 w/ 피해자 불일치(검색결과 X)':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=False)
            info_C = process_info(df_case, type='C', is_REF=True)
            info_V = process_info(df_case,
                                  type='V', is_REF=True)  # 피해자 대조
            write_alleles(info=info_C, hwp_control=hwp_control, num_slot=1,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])
            write_alleles(info=info_V, hwp_control=hwp_control, num_slot=2,
                          list_marker= self.ddi_present.combined_result.list_marker_ordered[:-3],
                          nickname='피해자')
            hwp_control.PutFieldText("num_evidence_result{{0}}", info_C['link_num_evidence'])
            hwp_control.PutFieldText("gender_result{{0}}", gender_to_text(info_C['profile']))
            hwp_control.PutFieldText("text_etc_locus{{0}}", info_C['str_etc'] + info_V['str_etc'])
        elif type_report == 'C 검출 w/ 피해자 불일치, ND(검색결과 X)':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=False)
            info_C = process_info(df_case, type='C', is_REF=True)
            info_ND = process_info(df_case, type='ND')
            info_V = process_info(df_case,
                                  type='V', is_REF=True)  # 피해자 대조
            write_alleles(info=info_C, hwp_control=hwp_control, num_slot=1,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])
            write_alleles(info=info_ND, hwp_control=hwp_control, num_slot=2,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])
            write_alleles(info=info_V, hwp_control=hwp_control, num_slot=3,
                          list_marker= self.ddi_present.combined_result.list_marker_ordered[:-3],
                          nickname='피해자')
            hwp_control.PutFieldText("num_evidence_result{{0}}", info_C['link_num_evidence'])
            hwp_control.PutFieldText("num_evidence_ND_result{{0}}", info_ND['link_num_evidence'])
            hwp_control.PutFieldText("gender_result{{0}}", gender_to_text(info_C['profile']))
            hwp_control.PutFieldText("text_etc_locus{{0}}", info_C['str_etc'])
            hwp_control.PutFieldText("num_evidence_ND_result{{0}}", info_ND['link_num_evidence'])
            hwp_control.PutFieldText("text_etc_locus{{0}}", 'ND : 디엔에이형이 검출되지 않음.\r' + info_C['str_etc'] + info_V['str_etc'])
        elif type_report == 'C 검출 w/ 피해자 일치(검색결과 X)':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=False)
            info_C = process_info(df_case, type='C', is_REF=True)
            info_V = process_info(df_case,
                                  type='V', is_REF=True)  # 피해자 대조
            info_v = process_info(df_case,
                                  type='v', is_REF=False, REF_type='V')  # 피해자 일치

            write_alleles(info=info_C, hwp_control=hwp_control, num_slot=1,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])
            write_alleles(info=info_v, hwp_control=hwp_control, num_slot=2,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[
                                      :-3])  # 부검이 아닌 REF는 SE33, PENTA_D, PENTA_E 제외
            write_alleles(info=info_V, hwp_control=hwp_control, num_slot=3,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3],
                          nickname='피해자')
            hwp_control.PutFieldText("num_evidence_result{{0}}", info_C['link_num_evidence'])
            hwp_control.PutFieldText("gender_result{{0}}", gender_to_text(info_C['profile']))
            hwp_control.PutFieldText("num_evidence_v_result{{0}}", info_v['link_num_evidence'])
            hwp_control.PutFieldText("float_mp_report{{0}}", info_V['mp'][:3]) # Matching Probalbility의 소수 부분
            hwp_control.PutFieldText("exp_mp_report{{0}}", info_V['mp'].split('+')[1]) #Matching Probability의 지수 부분
            hwp_control.PutFieldText("text_etc_locus{{0}}", info_C['str_etc'] + info_V['str_etc'])
        elif type_report == 'C 검출 w/ 피해자 일치, ND(검색결과 X)':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=False)
            info_C = process_info(df_case, type='C', is_REF=True)
            info_V = process_info(df_case,
                                  type='V', is_REF=True)  # 피해자 대조
            info_v = process_info(df_case,
                                  type='v', is_REF=False, REF_type='V')  # 피해자 일치
            info_ND = process_info(df_case, type='ND')
            write_alleles(info=info_C, hwp_control=hwp_control, num_slot=1,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])
            write_alleles(info=info_v, hwp_control=hwp_control, num_slot=2,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[
                                      :-3])  # 부검이 아닌 REF는 SE33, PENTA_D, PENTA_E 제외
            write_alleles(info=info_ND, hwp_control=hwp_control, num_slot=3,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])
            write_alleles(info=info_V, hwp_control=hwp_control, num_slot=4,
                          list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3],
                          nickname='피해자')
            hwp_control.PutFieldText("num_evidence_result{{0}}", info_C['link_num_evidence'])
            hwp_control.PutFieldText("gender_result{{0}}", gender_to_text(info_C['profile']))
            hwp_control.PutFieldText("num_evidence_v_result{{0}}", info_v['link_num_evidence'])
            hwp_control.PutFieldText("float_mp_report{{0}}", info_V['mp'][:3]) # Matching Probalbility의 소수 부분
            hwp_control.PutFieldText("exp_mp_report{{0}}", info_V['mp'].split('+')[1]) #Matching Probability의 지수 부분
            hwp_control.PutFieldText("num_evidence_ND_result{{0}}", info_ND['link_num_evidence'])
            hwp_control.PutFieldText("text_etc_locus{{0}}", 'ND : 디엔에이형이 검출되지 않음.\r' + info_C['str_etc'] + info_V['str_etc'])
        elif type_report == 'Complicate':
            generate_file(hwp_control=hwp_control, path_form=self.root + self.path_form_report[type_report],
                          filename=filename_new, y23=True)
            str_etc = ""
            str_etc_y23 = ""
            cnt_slot = 1
            list_types = list(set(df_case['DB Type 1'].values))
            
            for type in list_types:
                if type=='ND':
                    info_target = process_info(df_case, type='ND')
                    write_alleles(info=info_target, hwp_control=hwp_control, num_slot=cnt_slot,
                                  list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])
                elif type.isupper():  # type이 대문자 = REF 샘플
                    info_target = process_info(df_case, type=type, is_REF=True)
                    write_alleles(info=info_target, hwp_control=hwp_control, num_slot=cnt_slot,
                                  list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])
                else:   # type이 소문자 = REF 일치건
                    info_target = process_info(df_case, type=type, is_REF=False, REF_type=type.upper())
                    write_alleles(info=info_target, hwp_control=hwp_control, num_slot=cnt_slot,
                                  list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3])
                cnt_slot+=1
                str_etc = str_etc + info_target['str_etc']

            list_types_Y = list(set(df_case['Y Type'].values))
            cnt_slot=1
            if '' in list_types_Y: list_types_Y.remove('')
            if (list_types_Y):  # 입력할 Y23 Data가 있으면...
                for type in list_types_Y:
                    if type == 'ND':
                        info_ND = process_info(df_case, type='ND', is_y=True)
                        write_alleles(info=info_ND, hwp_control=hwp_control, num_slot=cnt_slot,
                                      list_marker=self.ddi_present.combined_result.list_marker_ordered[:-3], y23=True)
                    elif type.isupper():  # type이 대문자 = REF 샘플
                        info_target = process_info(df_case, type=type, is_REF=True, is_y=True)
                        write_alleles(info=info_target, hwp_control=hwp_control, num_slot=cnt_slot,
                                      list_marker=self.ddi_present.combined_result_y23.list_marker_ordered, y23=True)
                    else:  # type이 소문자 = REF 일치건
                        info_target = process_info(df_case, type=type, is_REF=False, is_y=True, REF_type=type.upper())
                        write_alleles(info=info_target, hwp_control=hwp_control, num_slot=cnt_slot,
                                      list_marker=self.ddi_present.combined_result_y23.list_marker_ordered, y23=True)
                    cnt_slot += 1
                    str_etc_y23 = str_etc_y23 + info_target['str_etc']
                hwp_control.PutFieldText("text_etc_locus_y23{{0}}", str_etc_y23)
            # hwp_control.PutFieldText("text_etc_result{{0}}", '1) 개인식별지수란 감정물의 디엔에이가 동일인으로부터 '
            #                                                  '유래되어서 디엔에이형이 일치할 확률 대 다른 사람으로부터'
            #                                                  ' 유래되었으나 우연히 디엔에이형이 일치할 확률의 비임.\r'
            #                                                  '2)「디엔에이신원확인정보의 이용 및 보호에 관한 법률」에 따라, '
            #                                                  '신원이 확인된 본 건 관련 범죄현장 증거물의 디엔에이형은'
            #                                                  ' 데이터베이스에서 삭제하겠음.\r'
            #                                                  '3) Y-STR 디엔에이형이 일치할 경우, 동일부계 남성이 배제되지 않음.\r'
            #                                                  '4) 감정물은 실험에 전량 소모하였음.')
            hwp_control.PutFieldText("text_etc_locus{{0}}", str_etc)
        hwp_control.Run("MoveDocBegin")
        hwp_control.Save()



    # hwnd = win32gui.FindWindow(None, '빈 문서 1 - 한글') # 한/글 창의 윈도우핸들값을 알아내서
    # win32gui.ShowWindow(hwnd,0) # 한/글 창을 백그라운드로 숨김
    # field_list = [i for i in hwp_control.GetFieldList().split("\x02")]
    # for field in field_list:
    # hwp_control.MoveToField(f'{field}{{{{{page}}}}}')    #커서를 해당 누름틀로 이동(작성과정을 지켜보기 위함, 없어도 무관), {{{{{page}}}}} 원하는 페이지에 access하기 위해선 {{1}}대신 앞의 변수를 사용

    # Data tab
    def click_btn_export_barcode(self):
        """ddi_present의 df_evidence의 데이터를 form_barcode.xls에 복사한다"""
        df_total = self.ddi_present.df_evidence[self.ddi_present.df_evidence['분류'] != 'Unassigned'] # 실험에 사용되지 않은 샘플을 제거한 데이터프레임 생성
        df_total.reset_index(inplace = True)
        wb_form = load_workbook(self.root + '/Form/form_barcode.xlsm', read_only=False, keep_vba=True)
        ws_form = wb_form.active
        row_start = 2
        for idx, line in df_total.iterrows():
            ws_form.cell(row=row_start + idx, column=3).value = line['증거물번호']
            ws_form.cell(row=row_start + idx, column=4).value = line['의뢰관서']
            ws_form.cell(row=row_start + idx, column=5).value = line['의뢰지역']
            ws_form.cell(row=row_start + idx, column=6).value = line['문서번호']
            ws_form.cell(row=row_start + idx, column=7).value = line['시행일자']
            ws_form.cell(row=row_start + idx, column=8).value = line['사건관련자']
            ws_form.cell(row=row_start + idx, column=9).value = line['접수번호']
            ws_form.cell(row=row_start + idx, column=10).value = line['접수일자']
            ws_form.cell(row=row_start + idx, column=11).value = line['담당자']
            ws_form.cell(row=row_start + idx, column=12).value = line['감정물']
        filename = self.ddi_present.date + '-' + self.ddi_present.analyst + '-barcode'
        ext = ".xlsm"
        wb_form.save(self.ddi_present.location_save + '/ETC/' + filename + ext)
        self.open_xls_file(self.ddi_present.location_save + '/ETC/' + filename + ext)
        QMessageBox.information(self, "Notice", "Work complete.")

    def click_btn_onsite_request(self):
        """업무분장 NFIS 파일을 입력받아 소내의뢰 시트와 증거물에 붙힐 라벨을 생성한다"""
        dict_evidenceType = {"약성분 분석" : "약독물실",
                             "생체시료중 기타 마약류 분석" : "약독물실",
                             "일반독물 분석" : "약독물실",
                             "혈중알코올농도" : "분석화학실",
                             "콘돔성분검사" : "분석화학실",
                             "음주 대사체 분석" : "분석화학실",
                             "화공약품(유해화학물질)" : "분석화학실",
                             "인화성액체" : "분석화학실",
                             "착화탄, 연소잔류물" : "분석화학실",
                             "압수품중 대마 분석" : "약독물실",
                             "생체시료중 환각물질 분석" : "약독물실",
                             "화학정밀정량분석" : "분석화학실",
                             "모발중 메트암페타민류 분석" : "약독물실"}
        filename = self.import_file(extension='xls(*.xls)', copy_needed=True)
        # 업무분장 NFIS파일이 xls이므로 openpyxl 사용을 위해 xlsx파일로 전환
        wb = self.dispatch_excel.Workbooks.Open(filename)
        wb.SaveAs(os.path.realpath(filename + 'x'), FileFormat=51)  # 51 : xlsx 확장자
        wb.Close()
        filename = filename + 'x'
        df_onsite = self.xls_to_dataframe(file_input=filename, column=True)
        df_onsite = df_onsite[df_onsite["처리실(처리자)"] != "본인"]
        def abbreviate(str_input):  #줄임말 처리를 위한 미니함수
            dict_evidenceAbbreviation = {"혈액": "혈액",
                                         "소변": "소변",
                                         "생식기": "질액",
                                         "슬라이드": "질액",
                                         "면봉": "질액",
                                         "생체시료중 기타 마약류 분석": "마약류 분석",
                                         "혈중알코올농도": "혈중알콜"}
            for key in dict_evidenceAbbreviation.keys():
                if str_input.find(key) != -1:
                    return dict_evidenceAbbreviation[key]
            return str_input    # 줄임말이 없으면 그대로 반환
        df_onsite["evidenceName"] = df_onsite["감정물-감정유형"].apply(lambda x:x.split(':')[1].split('-')[0])
        df_onsite["evidenceType"] = df_onsite["감정물-감정유형"].apply(lambda x:x.split(':')[1].split('-')[1])
        df_onsite["처리실(처리자)"] = df_onsite["감정물-감정유형"].apply(lambda x:dict_evidenceType[x.split(':')[1].split('-')[1]])
        # 소내의뢰 시트 및 라벨 생성
        wb = load_workbook(self.root + '/Form/form_onsiterequest.xlsx', data_only=True, read_only=False)
        ws_sheet = wb["sheet"]
        idx_sheet = 0
        idx_label = 0
        groupby_division = df_onsite.groupby("처리실(처리자)")
        for division, group in groupby_division:
            groupby_number = group.groupby("접수번호")
            for number, groupling in groupby_number:
                ws_sheet.cell(row=idx_sheet + 2, column=2).value = groupling['의뢰관서'].unique()[0]    # 의뢰관서
                ws_sheet.cell(row=idx_sheet + 2, column=3).value = number   # 접수번호
                list_abbreviatedEvidenceName = list(map(abbreviate, groupling['evidenceName'].unique()))
                ws_sheet.cell(row=idx_sheet + 2, column=4).value = ", ".join(list_abbreviatedEvidenceName)    # 감정물
                ws_sheet.cell(row=idx_sheet + 2, column=5).value = ", ".join(map(abbreviate, groupling['evidenceType'].unique()))   # 의뢰내용
                ws_sheet.cell(row=idx_sheet + 2, column=6).value = division # 의뢰분서
                ws_sheet.cell(row=idx_sheet + 2, column=7).value = time.strftime('%Y-%m-%d', time.localtime(time.time()))
                idx_sheet = idx_sheet + 1
                for type in list_abbreviatedEvidenceName:   # Label 생성
                    wb['label'].cell(row=int(idx_label/3) + 1, column = idx_label%3 + 1).value = number + '\n' + type
                    wb['label'].cell(row=int(idx_label / 3) + 1, column=idx_label % 3 + 1).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                    idx_label = idx_label + 1
        filename = self.ddi_present.date + '-' + self.ddi_present.analyst + '-onsiteRequest'
        ext = ".xlsx"
        wb.save(self.ddi_present.location_save + '/ETC/' + filename + ext)
        wb.close()
        self.open_xls_file(self.ddi_present.location_save + '/ETC/' + filename + ext)
        QMessageBox.information(self, "Notice", "Work complete.")

def except_hook(cls, exception, traceback): # for PyQt5.5 debugging
    sys.__excepthook__(cls, exception, traceback)


if __name__ == "__main__":
    Main_app = QApplication(sys.argv)
    GUI_EntryForm = EntryForm()
    GUI_EntryForm.setFixedSize(GUI_EntryForm.size())
    GUI_EntryForm.show()
    sys.excepthook = except_hook    # for PyQt5.5 debugging
    sys.exit(Main_app.exec_())
